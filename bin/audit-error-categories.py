#!/usr/bin/env python3
"""Audit error categories in session_costs over a rolling window.

Phase 0 of the autonomous ❌-Watcher Managed Agent design
(``docs/proposals/watcher-design-20260521-210800.md``). Pulls every
``session_costs`` row with ``outcome='error'`` over the last N days,
joins to ``investigations`` for ``error_message``, normalizes the
error string by stripping volatile identifiers (session IDs, event IDs,
inv IDs, file paths, commit SHAs, filename timestamps), buckets the
results by sha1[:16] of the normalized string, and ranks the top 10
categories.

Gates Phase 1 of the watcher rollout:
    top-3 share ≥ 70%           → proceed to Phase 1
    top-3 share 50-70%          → expand taxonomy to top-6 for diagnose-only path
    "unknown_terminalization" > 50%
                                → STOP, redesign as taxonomy parser first

Usage:
    python bin/audit-error-categories.py [--window-days 60] [--out PATH]
                                         [--portco acme] [--no-xlsx]

Sheets in the output workbook:
    1. ``Top categories``       — rank, hash, normalized message, count,
                                  share %, first_seen, last_seen, sample
                                  session_ids, sample inv_ids.
    2. ``Decision``             — total errors, top-3 share, decision rule
                                  verdict, distinct categories, max
                                  distinct hashes per category (Phase 0.5
                                  pre-check).
    3. ``Raw samples``          — first 200 raw rows with session_id,
                                  recorded_at, raw_error_message,
                                  normalized_message, error_message_hash.

Exit codes:
    0  — success (xlsx written, decision printed).
    1  — internal error (caught exception; details in stderr).
    2  — Postgres unreachable.
    3  — no error rows in the window (still writes empty xlsx).
"""

from __future__ import annotations

import argparse
import hashlib
import logging
import os
import re
import sys
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
ORCH_DIR = REPO_ROOT / "orchestrator"
sys.path.insert(0, str(ORCH_DIR))


log = logging.getLogger("audit_error_categories")


# ───────────────────────────────────────────────────────────────────────
# Setup helpers
# ───────────────────────────────────────────────────────────────────────


def _load_env() -> None:
    """Manual dotenv loader matching ``orchestrator/config.py``."""
    dotenv = REPO_ROOT / ".env"
    if not dotenv.exists():
        return
    for line in dotenv.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip())


def _setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )


# ───────────────────────────────────────────────────────────────────────
# Data access
# ───────────────────────────────────────────────────────────────────────


def _connect_pg():
    import psycopg2  # noqa: WPS433

    db_url = os.environ.get("DATABASE_URL", "")
    if not db_url:
        raise RuntimeError(
            "DATABASE_URL is unset. Set it locally via .env or on Railway."
        )
    return psycopg2.connect(db_url)


def _raw_usage_column_exists(conn) -> bool:
    """``session_costs.raw_usage_json`` is optional (Plan #35 spec line 91)."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = 'session_costs'
              AND column_name = 'raw_usage_json'
            LIMIT 1
            """
        )
        return cur.fetchone() is not None


def _fetch_error_rows(
    conn, *, days: int, portco_key: str | None
) -> list[dict]:
    """Return one row per failed investigation in the window.

    Source-of-truth is ``investigations`` (immutable append-only failure
    log), NOT ``session_costs`` (mutable per-session ledger where
    ``outcome`` is upserted and a successful follow-up will erase a
    prior failure on the same session_id). Time-bucket by
    ``investigations.completed_at`` so each failure is counted exactly
    once.

    ``session_costs`` is joined LATERAL for attribution only — agent_id,
    trigger, portco_key, and (optionally) ``raw_usage_json`` fall-back
    for the error message. The cost row is one per session_id, so even
    on a reused session the join is well-defined.
    """
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    has_raw = _raw_usage_column_exists(conn)
    raw_select = (
        "sc.raw_usage_json::TEXT AS raw_usage_text" if has_raw else "NULL AS raw_usage_text"
    )
    portco_clause = "AND i.portco_key = %s" if portco_key else ""
    sql = f"""
        SELECT
            i.session_id,
            COALESCE(i.completed_at, i.started_at) AS recorded_at,
            sc.agent_id,
            sc.trigger,
            i.portco_key,
            i.id AS inv_id,
            i.error_message,
            i.status AS inv_status,
            {raw_select}
        FROM investigations i
        LEFT JOIN session_costs sc
            ON sc.session_id = i.session_id
        WHERE i.status = 'failed'
          AND COALESCE(i.completed_at, i.started_at) >= %s
          {portco_clause}
        ORDER BY COALESCE(i.completed_at, i.started_at) DESC
    """
    params: tuple
    if portco_key:
        params = (cutoff, portco_key)
    else:
        params = (cutoff,)
    with conn.cursor() as cur:
        cur.execute(sql, params)
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
    return [dict(zip(cols, r)) for r in rows]


# ───────────────────────────────────────────────────────────────────────
# Normalization
# ───────────────────────────────────────────────────────────────────────


# Volatile identifier patterns the watcher's hash must collapse so that
# same-root errors land in the same bucket. Order matters — longer / more
# specific patterns first so a generic ``[0-9a-f]{7,40}`` SHA strip does
# not eat the prefix of a labeled ID like ``session_abc1234567``.
_NORMALIZATION_PATTERNS: tuple[tuple[re.Pattern, str], ...] = (
    # Labeled IDs (Anthropic SDK shapes)
    # ``sesn_`` is the real Managed Agents session ID prefix used in this
    # repo (see orchestrator/ref_footer_test.py); ``session_`` covers
    # the legacy Anthropic SDK shape that appears in older log lines.
    (re.compile(r"sesn_[A-Za-z0-9]+"), "sesn_<ID>"),
    (re.compile(r"session_[A-Za-z0-9]+"), "session_<ID>"),
    (re.compile(r"msg_[A-Za-z0-9]+"), "msg_<ID>"),
    (re.compile(r"event_[A-Za-z0-9]+"), "event_<ID>"),
    (re.compile(r"req_[A-Za-z0-9]+"), "req_<ID>"),
    (re.compile(r"memstore_[A-Za-z0-9]+"), "memstore_<ID>"),
    (re.compile(r"agent_[A-Za-z0-9]+"), "agent_<ID>"),
    (re.compile(r"toolu_[A-Za-z0-9]+"), "toolu_<ID>"),
    # Investigation IDs in error messages (e.g. "inv_id=12345", "inv 67")
    (re.compile(r"inv_id=\d+"), "inv_id=<N>"),
    (re.compile(r"\binv \d+\b"), "inv <N>"),
    # Slack channel + ts shapes
    (re.compile(r"\bC[A-Z0-9]{8,}\b"), "C<CHANNEL>"),
    (re.compile(r"\b\d{10}\.\d{6}\b"), "<SLACK_TS>"),
    # Filename timestamps: 20260521-153045 or 2026-05-21T15:30:45
    (re.compile(r"\d{8}-\d{6}"), "<FNTS>"),
    (re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:[Z+-]\S*)?"),
     "<ISO_TS>"),
    # File paths (absolute or POSIX-style relative with extension)
    (re.compile(r"(?:/[\w.-]+)+\.[a-zA-Z0-9]{1,6}"), "<PATH>"),
    # UUIDs — MUST come before the bare-SHA strip below; otherwise the
    # \b[0-9a-f]{7,40}\b SHA regex eats each hex group of a UUID
    # individually and the UUID pattern never matches what is left.
    (re.compile(
        r"\b[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
        r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}\b"
    ), "<UUID>"),
    # Git commit SHAs (7-40 hex chars, lowercase) — after labeled IDs + UUIDs
    (re.compile(r"\b[0-9a-f]{7,40}\b"), "<SHA>"),
    # Bare integers >= 4 digits (line numbers, sizes, counts)
    (re.compile(r"\b\d{4,}\b"), "<N>"),
    # Squeeze runs of whitespace to single space
    (re.compile(r"\s+"), " "),
)


def _first_line(text: str | None) -> str:
    if not text:
        return ""
    line = text.strip().split("\n", 1)[0]
    # Truncate absurdly long single lines (stack-traced messages do this)
    return line[:500]


def _normalize_error(message: str | None) -> str:
    """Apply the volatile-identifier strip set to a raw error message."""
    out = _first_line(message)
    for pattern, replacement in _NORMALIZATION_PATTERNS:
        out = pattern.sub(replacement, out)
    return out.strip()


def _hash_normalized(normalized: str) -> str:
    return hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:16]


def _resolve_error_text(row: dict) -> str:
    """Extract the best-available error text from the row.

    Preference order:
        1. ``investigations.error_message`` (set by terminalize_lifecycle)
        2. ``session_costs.raw_usage_json`` → ``error.message`` field
        3. Sentinel ``unknown_terminalization`` when both are absent.
    """
    msg = row.get("error_message")
    if msg:
        return msg
    raw = row.get("raw_usage_text")
    if raw:
        # Parse as JSON. The cheap-string-probe alternative truncated on
        # escaped quotes (e.g. ``"Invalid \"Field\" value"``), which
        # silently collapsed unrelated validation failures into one hash.
        try:
            import json  # local import — only the fallback path needs it

            obj = json.loads(raw)
            if isinstance(obj, dict):
                err = obj.get("error")
                if isinstance(err, dict):
                    m = err.get("message")
                    if isinstance(m, str) and m:
                        return m
                # Some legacy rows used a flat error_summary field.
                summary = obj.get("error_summary")
                if isinstance(summary, str) and summary:
                    return summary
        except (ValueError, TypeError):
            pass  # fall through to the sentinel
    return "unknown_terminalization"


# ───────────────────────────────────────────────────────────────────────
# Aggregation
# ───────────────────────────────────────────────────────────────────────


def _bucket_errors(rows: list[dict]) -> dict[str, dict]:
    """Group rows by ``error_message_hash``.

    Returns ``{hash: {normalized, count, first_seen, last_seen,
    sample_session_ids, sample_inv_ids, raw_variants}}``.
    """
    buckets: dict[str, dict] = {}
    for r in rows:
        raw = _resolve_error_text(r)
        normalized = _normalize_error(raw)
        h = _hash_normalized(normalized)
        b = buckets.setdefault(
            h,
            {
                "normalized": normalized,
                "count": 0,
                "first_seen": r["recorded_at"],
                "last_seen": r["recorded_at"],
                "sample_session_ids": [],
                "sample_inv_ids": [],
                "raw_variants": set(),
            },
        )
        b["count"] += 1
        ts = r["recorded_at"]
        if ts < b["first_seen"]:
            b["first_seen"] = ts
        if ts > b["last_seen"]:
            b["last_seen"] = ts
        if len(b["sample_session_ids"]) < 5:
            b["sample_session_ids"].append(r["session_id"])
        if r.get("inv_id") and len(b["sample_inv_ids"]) < 5:
            b["sample_inv_ids"].append(r["inv_id"])
        # Track up to 5 distinct raw first-lines per bucket so Phase 0.5
        # can spot under-normalization (same root → multiple hashes).
        if len(b["raw_variants"]) < 5:
            b["raw_variants"].add(_first_line(raw))
    return buckets


def _decision_verdict(
    total_errors: int,
    sorted_buckets: list[tuple[str, dict]],
) -> tuple[str, dict]:
    """Apply the Phase 0 gating rule. Returns (verdict, metrics)."""
    if total_errors == 0:
        return ("NO_DATA", {"reason": "no error rows in window"})
    counts = [b["count"] for _, b in sorted_buckets]
    top_n = lambda n: sum(counts[:n])  # noqa: E731 — short-form helper
    top3 = top_n(3)
    top3_share = top3 / total_errors

    # Locate the unknown_terminalization bucket if present.
    unknown_share = 0.0
    for _, b in sorted_buckets:
        if b["normalized"] == "unknown_terminalization":
            unknown_share = b["count"] / total_errors
            break

    metrics = {
        "total_errors": total_errors,
        "top3_count": top3,
        "top3_share_pct": round(100.0 * top3_share, 2),
        "unknown_terminalization_share_pct": round(100.0 * unknown_share, 2),
        "distinct_categories": len(sorted_buckets),
    }

    if unknown_share > 0.50:
        return (
            "STOP_REDESIGN",
            {**metrics, "reason": "unknown_terminalization > 50% — taxonomy parser needed first"},
        )
    if top3_share >= 0.70:
        return ("PROCEED_PHASE_1", {**metrics, "reason": "top-3 share ≥ 70%"})
    if top3_share >= 0.50:
        return (
            "EXPAND_TO_TOP_6",
            {**metrics, "reason": "top-3 share 50-70% — widen taxonomy to top-6 diagnose-only"},
        )
    return (
        "STOP_REDESIGN",
        {**metrics, "reason": "top-3 share < 50% — long-tail dominant, watcher wedge too narrow"},
    )


def _phase_05_precheck(buckets: dict[str, dict]) -> dict:
    """Phase 0.5 sanity: detect category FRAGMENTATION.

    Under-normalization shows up as one logical error type splitting
    across multiple hashes — e.g. ``TypeError: foo`` and ``TypeError: bar``
    landing in different buckets when the volatile bit was actually
    after the colon. The inverse signal (one hash with many distinct
    raw variants) is GOOD: it means normalization collapsed them
    correctly.

    We approximate "logical error type" by the first token of the
    normalized message (typically the exception class or a leading
    identifier). If a single token appears across more than 3 hashes,
    flag it for regex review.
    """
    by_token: dict[str, list[str]] = {}
    for h, b in buckets.items():
        normalized = b.get("normalized") or ""
        # Strip trailing punctuation so "TypeError:" and "TypeError"
        # share a token.
        token = (
            normalized.split(" ", 1)[0].rstrip(":,;.")
            if normalized
            else "(empty)"
        )
        by_token.setdefault(token, []).append(h)

    fragmentation = {t: len(hs) for t, hs in by_token.items() if len(hs) > 1}
    threshold = 3
    fragmented_over = {
        t: hs for t, hs in by_token.items() if len(hs) > threshold
    }
    max_frag = max(fragmentation.values()) if fragmentation else 1
    return {
        "max_hashes_per_token": max_frag,
        "fragmented_tokens": sorted(fragmented_over.keys()),
        "threshold": threshold,
    }


# ───────────────────────────────────────────────────────────────────────
# Workbook output
# ───────────────────────────────────────────────────────────────────────


def _write_xlsx(
    out_path: Path,
    *,
    sorted_buckets: list[tuple[str, dict]],
    total_errors: int,
    verdict: str,
    decision_metrics: dict,
    phase_05: dict,
    raw_rows: list[dict],
    days: int,
    portco_key: str | None,
) -> None:
    """Emit the 3-sheet workbook."""
    from openpyxl import Workbook  # noqa: WPS433

    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None  # openpyxl always creates a default sheet
    ws1.title = "Top categories"
    ws1.append(
        [
            f"Portco: {portco_key or '(all)'}",
            f"Window: last {days} days",
            f"Total errors: {total_errors}",
        ]
    )
    ws1.append([])
    ws1.append(
        [
            "Rank",
            "Hash",
            "Normalized message",
            "Count",
            "Share (%)",
            "First seen (UTC)",
            "Last seen (UTC)",
            "Sample session IDs",
            "Sample inv IDs",
            "Distinct raw variants",
        ]
    )
    for rank, (h, b) in enumerate(sorted_buckets[:10], start=1):
        share = (100.0 * b["count"] / total_errors) if total_errors else 0.0
        ws1.append(
            [
                rank,
                h,
                b["normalized"][:300],
                b["count"],
                round(share, 2),
                b["first_seen"].isoformat() if b["first_seen"] else "",
                b["last_seen"].isoformat() if b["last_seen"] else "",
                ", ".join(b["sample_session_ids"]),
                ", ".join(str(i) for i in b["sample_inv_ids"]),
                len(b["raw_variants"]),
            ]
        )

    ws2 = wb.create_sheet("Decision")
    ws2.append([f"Window: last {days} days", f"Portco: {portco_key or '(all)'}"])
    ws2.append([])
    ws2.append(["Metric", "Value"])
    ws2.append(["Verdict", verdict])
    for k, v in decision_metrics.items():
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["Phase 0.5 pre-check (fragmentation across hashes)", ""])
    ws2.append(["Max hashes sharing one first-token", phase_05["max_hashes_per_token"]])
    ws2.append(["Fragmentation threshold (target ≤)", phase_05.get("threshold", 3)])
    ws2.append(
        ["Tokens fragmented over threshold",
         ", ".join(phase_05["fragmented_tokens"]) or "(none)"]
    )

    ws3 = wb.create_sheet("Raw samples")
    ws3.append([f"First {min(200, len(raw_rows))} of {len(raw_rows)} error rows"])
    ws3.append([])
    ws3.append(
        [
            "recorded_at (UTC)",
            "session_id",
            "agent_id",
            "trigger",
            "inv_id",
            "inv_status",
            "raw error (first line)",
            "normalized",
            "hash",
        ]
    )
    for r in raw_rows[:200]:
        raw = _resolve_error_text(r)
        normalized = _normalize_error(raw)
        ws3.append(
            [
                r["recorded_at"].isoformat() if r["recorded_at"] else "",
                r["session_id"],
                r.get("agent_id") or "",
                r.get("trigger") or "",
                r.get("inv_id") or "",
                r.get("inv_status") or "",
                _first_line(raw),
                normalized[:300],
                _hash_normalized(normalized),
            ]
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


# ───────────────────────────────────────────────────────────────────────
# Main
# ───────────────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Audit error categories from session_costs over a rolling window. "
            "Phase 0 of the autonomous ❌-watcher rollout."
        )
    )
    parser.add_argument("--window-days", type=int, default=60)
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output xlsx path. Defaults to docs/audits/error-categories-<UTCstamp>.xlsx",
    )
    parser.add_argument(
        "--portco",
        type=str,
        default=None,
        help="Filter to a single portco_key. Default: all portcos.",
    )
    parser.add_argument(
        "--no-xlsx",
        action="store_true",
        help="Skip xlsx write — print decision to stdout only.",
    )
    args = parser.parse_args(argv)

    _setup_logging()
    _load_env()

    try:
        conn = _connect_pg()
    except Exception as exc:  # pragma: no cover — depends on local env
        log.error("Postgres unreachable: %s", exc)
        return 2

    try:
        rows = _fetch_error_rows(
            conn, days=args.window_days, portco_key=args.portco
        )
    except Exception:
        log.error("Query failed:\n%s", traceback.format_exc())
        conn.close()
        return 1
    finally:
        if not conn.closed:
            conn.close()

    total_errors = len(rows)
    buckets = _bucket_errors(rows)
    sorted_buckets = sorted(
        buckets.items(), key=lambda kv: kv[1]["count"], reverse=True
    )
    verdict, metrics = _decision_verdict(total_errors, sorted_buckets)
    phase_05 = _phase_05_precheck(buckets)

    # Stdout summary — the operator reads this even if xlsx is skipped.
    print(f"Window: last {args.window_days} days")
    print(f"Portco filter: {args.portco or '(all)'}")
    print(f"Total error rows: {total_errors}")
    print(f"Distinct categories: {len(sorted_buckets)}")
    print(f"VERDICT: {verdict}")
    for k, v in metrics.items():
        print(f"  {k}: {v}")
    print(
        f"Phase 0.5 pre-check: max_hashes_per_token={phase_05['max_hashes_per_token']} "
        f"(target ≤ {phase_05.get('threshold', 3)})"
    )
    if phase_05["fragmented_tokens"]:
        print(
            f"  Tokens fragmented over threshold: {', '.join(phase_05['fragmented_tokens'])}"
        )

    print("\nTop 10 categories:")
    for rank, (h, b) in enumerate(sorted_buckets[:10], start=1):
        share = (100.0 * b["count"] / total_errors) if total_errors else 0.0
        print(
            f"  {rank:>2}. [{h}] {b['count']:>4} ({share:5.2f}%) "
            f"{b['normalized'][:120]}"
        )

    if total_errors == 0:
        log.warning("No error rows in window — writing empty xlsx and exiting 3.")

    if not args.no_xlsx:
        if args.out is None:
            stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
            args.out = REPO_ROOT / "docs" / "audits" / f"error-categories-{stamp}.xlsx"
        _write_xlsx(
            args.out,
            sorted_buckets=sorted_buckets,
            total_errors=total_errors,
            verdict=verdict,
            decision_metrics=metrics,
            phase_05=phase_05,
            raw_rows=rows,
            days=args.window_days,
            portco_key=args.portco,
        )
        print(f"\nWrote: {args.out}")

    return 3 if total_errors == 0 else 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
