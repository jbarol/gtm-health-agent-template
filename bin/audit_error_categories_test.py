"""Tests for ``bin/audit-error-categories.py`` (Phase 0 of ❌-watcher).

Fully mocked — no real Postgres. The script has a hyphen in its
filename, so we load it by path via importlib.util.

Covered:
- Normalization regex set collapses volatile IDs.
- ``_resolve_error_text`` prefers investigations.error_message, falls
  back to raw_usage_json, finally to ``unknown_terminalization``.
- ``_decision_verdict`` returns PROCEED / EXPAND / STOP per the gate rule.
- ``_bucket_errors`` counts, tracks first/last seen, samples session IDs.
- ``main()`` writes a 3-sheet workbook and exits 0 on a populated query.
- Empty result → exit 3.
- Postgres unavailable → exit 2.
"""

from __future__ import annotations

import importlib.util
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPT_PATH = REPO_ROOT / "bin" / "audit-error-categories.py"


# Same env bootstrap as bin/measure_deploy_risk_test.py — orchestrator/config.py
# raises on missing keys at module load.
for _key, _value in {
    "ANTHROPIC_API_KEY": "sk-ant-test",
    "SLACK_BOT_TOKEN": "xoxb-test",
    "SLACK_APP_TOKEN": "xapp-test",
    "SLACK_CHANNEL_ID": "C0TEST",
    "ENVIRONMENT_ID": "env_test",
    "DREAM_AGENT_ID": "agent_test_dream",
    "COORDINATOR_ID": "agent_test_coord",
    "QUICK_AGENT_ID": "agent_test_quick",
    "METHODOLOGY_STORE_ID": "memstore_test_m",
    "HEALTH_STORE_ID": "memstore_test_h",
    "DATABASE_URL": "postgres://test/db",
}.items():
    os.environ.setdefault(_key, _value)


def _load_script_module():
    for p in (REPO_ROOT / "orchestrator", REPO_ROOT / "bin"):
        if str(p) not in sys.path:
            sys.path.insert(0, str(p))
    spec = importlib.util.spec_from_file_location("audit_error_categories", SCRIPT_PATH)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture()
def aec():
    return _load_script_module()


# ───────────────────────────────────────────────────────────────────────
# Normalization
# ───────────────────────────────────────────────────────────────────────


def test_normalize_strips_session_ids(aec):
    msg = "session_runner: session_abc123def456 raised TypeError"
    out = aec._normalize_error(msg)
    assert "session_abc123def456" not in out
    assert "session_<ID>" in out


def test_normalize_strips_sesn_EXAMPLE(aec):
    """Real Managed Agents IDs use ``sesn_`` prefix (see ref_footer_test.py)."""
    msg = "sesn_EXAMPLE failed mid-flight"
    out = aec._normalize_error(msg)
    assert "sesn_EXAMPLE" not in out
    assert "sesn_<ID>" in out


def test_normalize_uuid_before_sha_does_not_fragment(aec):
    """SHA strip must not eat UUID hex groups individually."""
    a = "Session 019e6a5b-831e-7fe0-abb0-91ee7a97cfaa failed"
    b = "Session aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee failed"
    # Two distinct UUIDs in identical contexts should normalize to the same string.
    assert aec._normalize_error(a) == aec._normalize_error(b)
    assert "<UUID>" in aec._normalize_error(a)


def test_normalize_strips_msg_event_req_ids(aec):
    msg = "request req_01HXYZ failed at msg_AbCdEf for event_99zz"
    out = aec._normalize_error(msg)
    assert "req_<ID>" in out
    assert "msg_<ID>" in out
    assert "event_<ID>" in out
    assert "req_01HXYZ" not in out


def test_normalize_strips_inv_ids(aec):
    msg = "Lifecycle terminalize failed for inv_id=12345 (inv 67 retry)"
    out = aec._normalize_error(msg)
    assert "inv_id=<N>" in out
    assert "inv <N>" in out
    assert "12345" not in out
    # "67" is a small int and not collapsed by the bare-int rule (\d{4,}),
    # so it's only stripped via the labeled "inv N" rule above.


def test_normalize_strips_paths_and_shas(aec):
    msg = "AttributeError at /Users/jb/repos/x/orchestrator/main.py line 4242 commit 595d717c7047ba"
    out = aec._normalize_error(msg)
    assert "/Users/jb" not in out
    assert "<PATH>" in out
    assert "595d717c7047ba" not in out
    assert "<SHA>" in out
    assert "4242" not in out  # \d{4,} rule
    assert "<N>" in out


def test_normalize_strips_filename_timestamps(aec):
    msg = "Snapshot 20260521-153045 conflicts with 2026-05-21T15:30:45Z"
    out = aec._normalize_error(msg)
    assert "20260521-153045" not in out
    assert "2026-05-21T15:30:45" not in out
    assert "<FNTS>" in out
    assert "<ISO_TS>" in out


def test_normalize_strips_slack_channel_and_ts(aec):
    msg = "Slack post failed in C09ABC1234 at 1716315045.123456"
    out = aec._normalize_error(msg)
    assert "C09ABC1234" not in out
    assert "1716315045.123456" not in out
    assert "C<CHANNEL>" in out
    assert "<SLACK_TS>" in out


def test_normalize_collapses_same_root_to_same_hash(aec):
    """Two errors with different volatile IDs but same root should hash equal."""
    a = "Lifecycle terminalize failed for inv_id=12345 (commit 595d717)"
    b = "Lifecycle terminalize failed for inv_id=99999 (commit abcdef0)"
    assert aec._hash_normalized(aec._normalize_error(a)) == aec._hash_normalized(
        aec._normalize_error(b)
    )


def test_normalize_handles_none_and_empty(aec):
    assert aec._normalize_error(None) == ""
    assert aec._normalize_error("") == ""
    assert aec._hash_normalized("")  # still produces a deterministic hash


# ───────────────────────────────────────────────────────────────────────
# Resolution
# ───────────────────────────────────────────────────────────────────────


def test_resolve_prefers_investigations_error_message(aec):
    row = {
        "error_message": "from investigations",
        "raw_usage_text": '{"error": {"message": "from raw_usage"}}',
    }
    assert aec._resolve_error_text(row) == "from investigations"


def test_resolve_falls_back_to_raw_usage(aec):
    row = {
        "error_message": None,
        "raw_usage_text": '{"foo": 1, "error": {"message": "raw_value"}}',
    }
    assert aec._resolve_error_text(row) == "raw_value"


def test_resolve_falls_back_to_unknown(aec):
    row = {"error_message": None, "raw_usage_text": None}
    assert aec._resolve_error_text(row) == "unknown_terminalization"


def test_resolve_handles_escaped_quotes_in_raw_usage(aec):
    """The naive string-probe truncated at the first escaped quote."""
    row = {
        "error_message": None,
        "raw_usage_text": '{"error": {"message": "Invalid \\"Field\\" value rejected"}}',
    }
    assert (
        aec._resolve_error_text(row) == 'Invalid "Field" value rejected'
    )


def test_resolve_handles_error_summary_legacy_field(aec):
    row = {
        "error_message": None,
        "raw_usage_text": '{"error_summary": "WRITING_AGENT_FALLTHROUGH"}',
    }
    assert aec._resolve_error_text(row) == "WRITING_AGENT_FALLTHROUGH"


def test_resolve_invalid_json_falls_through_to_unknown(aec):
    row = {"error_message": None, "raw_usage_text": "not valid json {"}
    assert aec._resolve_error_text(row) == "unknown_terminalization"


# test_fetch_query_prefers_failed_status_in_lateral was superseded by
# test_fetch_query_uses_investigations_as_source_of_truth — the LATERAL
# preference logic is gone now that investigations is the FROM table.


# ───────────────────────────────────────────────────────────────────────
# Bucketing
# ───────────────────────────────────────────────────────────────────────


def _now():
    return datetime.now(timezone.utc)


def test_bucket_counts_and_tracks_window(aec):
    base = _now()
    rows = [
        {
            "session_id": f"sess_{i}",
            "recorded_at": base - timedelta(minutes=i),
            "agent_id": "agent_x",
            "trigger": "slack-adhoc",
            "portco_key": "acme",
            "inv_id": 100 + i,
            "inv_status": "terminal_failure",
            "error_message": f"TypeError at inv_id={i}",
            "raw_usage_text": None,
        }
        for i in range(7)
    ]
    buckets = aec._bucket_errors(rows)
    assert len(buckets) == 1, "All 7 rows should collapse to one bucket"
    only = next(iter(buckets.values()))
    assert only["count"] == 7
    assert only["first_seen"] < only["last_seen"]
    assert len(only["sample_session_ids"]) == 5  # capped


def test_bucket_separates_distinct_roots(aec):
    base = _now()
    rows = [
        {
            "session_id": "s1",
            "recorded_at": base,
            "agent_id": "a",
            "trigger": "t",
            "portco_key": "p",
            "inv_id": 1,
            "inv_status": "x",
            "error_message": "TypeError: bad",
            "raw_usage_text": None,
        },
        {
            "session_id": "s2",
            "recorded_at": base,
            "agent_id": "a",
            "trigger": "t",
            "portco_key": "p",
            "inv_id": 2,
            "inv_status": "x",
            "error_message": "ConnectionError: dropped",
            "raw_usage_text": None,
        },
    ]
    buckets = aec._bucket_errors(rows)
    assert len(buckets) == 2


# ───────────────────────────────────────────────────────────────────────
# Decision verdict
# ───────────────────────────────────────────────────────────────────────


def test_decision_proceed_when_top3_ge_70(aec):
    sorted_buckets = [
        ("h1", {"count": 50, "normalized": "A"}),
        ("h2", {"count": 30, "normalized": "B"}),
        ("h3", {"count": 5, "normalized": "C"}),
        ("h4", {"count": 15, "normalized": "D"}),
    ]
    verdict, _ = aec._decision_verdict(100, sorted_buckets)
    assert verdict == "PROCEED_PHASE_1"


def test_decision_expand_when_top3_50_to_70(aec):
    sorted_buckets = [
        ("h1", {"count": 25, "normalized": "A"}),
        ("h2", {"count": 20, "normalized": "B"}),
        ("h3", {"count": 15, "normalized": "C"}),
        ("h4", {"count": 40, "normalized": "D"}),
    ]
    verdict, _ = aec._decision_verdict(100, sorted_buckets)
    assert verdict == "EXPAND_TO_TOP_6"


def test_decision_stop_when_top3_below_50(aec):
    sorted_buckets = [
        ("h1", {"count": 15, "normalized": "A"}),
        ("h2", {"count": 12, "normalized": "B"}),
        ("h3", {"count": 10, "normalized": "C"}),
        ("h4", {"count": 63, "normalized": "D"}),
    ]
    verdict, _ = aec._decision_verdict(100, sorted_buckets)
    assert verdict == "STOP_REDESIGN"


def test_decision_stop_when_unknown_terminalization_gt_50(aec):
    sorted_buckets = [
        ("h1", {"count": 60, "normalized": "unknown_terminalization"}),
        ("h2", {"count": 30, "normalized": "B"}),
        ("h3", {"count": 10, "normalized": "C"}),
    ]
    verdict, metrics = aec._decision_verdict(100, sorted_buckets)
    assert verdict == "STOP_REDESIGN"
    assert "unknown_terminalization" in metrics["reason"]


def test_decision_no_data(aec):
    verdict, _ = aec._decision_verdict(0, [])
    assert verdict == "NO_DATA"


# ───────────────────────────────────────────────────────────────────────
# Phase 0.5 pre-check
# ───────────────────────────────────────────────────────────────────────


def test_phase_05_precheck_flags_fragmented_token(aec):
    """One leading token across 4 hashes → fragmentation flagged."""
    buckets = {
        "h1": {"normalized": "TypeError: bad type"},
        "h2": {"normalized": "TypeError: cannot subscript"},
        "h3": {"normalized": "TypeError: not iterable"},
        "h4": {"normalized": "TypeError: missing argument"},
        "h5": {"normalized": "ConnectionError: dropped"},
    }
    result = aec._phase_05_precheck(buckets)
    assert result["max_hashes_per_token"] == 4  # TypeError appears in 4 hashes
    assert "TypeError" in result["fragmented_tokens"]
    assert "ConnectionError" not in result["fragmented_tokens"]


def test_phase_05_precheck_collapses_volatile_not_flagged(aec):
    """A single hash with many raw variants is NORMAL — should NOT flag."""
    buckets = {
        "h1": {"normalized": "TypeError: foo"},  # raw_variants had 5 entries but they all hashed to h1
        "h2": {"normalized": "ConnectionError: dropped"},
    }
    result = aec._phase_05_precheck(buckets)
    assert result["fragmented_tokens"] == []
    assert result["max_hashes_per_token"] == 1


def test_phase_05_precheck_handles_empty_normalized(aec):
    buckets = {"h1": {"normalized": ""}, "h2": {"normalized": None}}
    result = aec._phase_05_precheck(buckets)
    # Two empty-normalized hashes share the "(empty)" token — flagged at 2,
    # but threshold is 3, so not "fragmented_over".
    assert result["fragmented_tokens"] == []


# ───────────────────────────────────────────────────────────────────────
# End-to-end main()
# ───────────────────────────────────────────────────────────────────────


def _build_mock_conn(rows: list[tuple]) -> MagicMock:
    """Build a psycopg2-shaped MagicMock; first .cursor() is the
    raw_usage_column probe, second is the data SELECT.
    """
    conn = MagicMock()
    conn.closed = False
    raw_cur = MagicMock()
    raw_cur.fetchone.return_value = None
    raw_cur.__enter__ = MagicMock(return_value=raw_cur)
    raw_cur.__exit__ = MagicMock(return_value=None)
    data_cur = MagicMock()
    data_cur.fetchall.return_value = rows
    data_cur.description = [
        ("session_id",), ("recorded_at",), ("agent_id",), ("trigger",),
        ("portco_key",), ("inv_id",), ("error_message",), ("inv_status",),
        ("raw_usage_text",),
    ]
    data_cur.__enter__ = MagicMock(return_value=data_cur)
    data_cur.__exit__ = MagicMock(return_value=None)
    conn.cursor.side_effect = [raw_cur, data_cur]
    return conn


def test_main_writes_xlsx_and_exits_0(aec, tmp_path):
    base = _now()
    rows = [
        ("s1", base, "a1", "slack-adhoc", "p1", 1, "TypeError: foo", "failed", None),
        ("s2", base - timedelta(minutes=1), "a1", "slack-adhoc", "p1", 2, "TypeError: foo", "failed", None),
        ("s3", base - timedelta(minutes=2), "a1", "slack-adhoc", "p1", 3, "ConnectionError: bar", "failed", None),
    ]
    conn = _build_mock_conn(rows)

    out = tmp_path / "audit.xlsx"
    with patch.object(aec, "_connect_pg", return_value=conn):
        rc = aec.main(["--window-days", "30", "--out", str(out)])
    assert rc == 0
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    assert "Top categories" in wb.sheetnames
    assert "Decision" in wb.sheetnames
    assert "Raw samples" in wb.sheetnames


def test_main_no_rows_exits_3(aec, tmp_path):
    conn = _build_mock_conn([])

    out = tmp_path / "audit-empty.xlsx"
    with patch.object(aec, "_connect_pg", return_value=conn):
        rc = aec.main(["--window-days", "60", "--out", str(out)])
    assert rc == 3
    assert out.exists()  # empty xlsx still written


def test_main_pg_unavailable_exits_2(aec, tmp_path):
    with patch.object(aec, "_connect_pg", side_effect=RuntimeError("no db")):
        rc = aec.main(["--window-days", "60", "--out", str(tmp_path / "x.xlsx")])
    assert rc == 2


def test_fetch_query_uses_investigations_as_source_of_truth(aec):
    """The audit must read failed investigations (immutable log), not
    session_costs (mutable per-session ledger where outcome upsert can
    erase a prior failure on the same session_id)."""
    raw_cur = MagicMock()
    raw_cur.fetchone.return_value = None
    raw_cur.__enter__ = MagicMock(return_value=raw_cur)
    raw_cur.__exit__ = MagicMock(return_value=None)
    data_cur = MagicMock()
    data_cur.fetchall.return_value = []
    data_cur.description = [("session_id",)]
    data_cur.__enter__ = MagicMock(return_value=data_cur)
    data_cur.__exit__ = MagicMock(return_value=None)
    conn = MagicMock()
    conn.cursor.side_effect = [raw_cur, data_cur]
    conn.closed = False

    aec._fetch_error_rows(conn, days=10, portco_key=None)

    sql_passed = data_cur.execute.call_args.args[0]
    # FROM clause must lead with investigations, not session_costs
    assert "FROM investigations i" in sql_passed
    # Filter must be on the persisted failed status
    assert "i.status = 'failed'" in sql_passed
    # Time bucket must use completed_at (terminal timestamp), not recorded_at
    assert "completed_at" in sql_passed
    # session_costs is joined LATERAL or LEFT JOIN for attribution only
    assert "LEFT JOIN session_costs" in sql_passed
    # Outcome column on session_costs must NOT be filtered (it's mutable)
    assert "sc.outcome" not in sql_passed


def test_main_no_xlsx_flag(aec):
    base = _now()
    conn = _build_mock_conn(
        [("s1", base, "a", "t", "p", 1, "TypeError", "failed", None)]
    )

    with patch.object(aec, "_connect_pg", return_value=conn):
        rc = aec.main(["--no-xlsx", "--window-days", "1"])
    assert rc == 0
