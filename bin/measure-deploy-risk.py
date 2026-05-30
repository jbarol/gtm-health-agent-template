#!/usr/bin/env python3
"""Measure deploy risk for Acme GTM Health Agent.

Plan #42 PR1 (D11, D17). Idempotent. Reads ``session_costs`` for the last
N days and emits a 3-sheet .xlsx histogram + a Slack admin DM with the
file attached. The output drives the data-first refinement of the
business-hours window (D2) and the "do we ship plan #43" decision (D17).

Runs locally (``python bin/measure-deploy-risk.py``) and on a monthly
GitHub Actions cron (``.github/workflows/measure-deploy-risk.yml``).

Usage:
    python bin/measure-deploy-risk.py [--days 60] [--out PATH] [--dry-run]
                                      [--portco acme]

Sheets in the output workbook:
    1. ``Sessions by hour``       — total session count per hour-of-day (0-23)
                                    from ``session_costs.recorded_at``.
    2. ``Error rate by hour``     — error count + error rate per hour from
                                    ``session_costs.outcome`` (added in
                                    migration ``00AB_session_costs_outcome.sql``).
    3. ``Deploys vs incidents``   — per-day rollup joining ``git log main``
                                    commit timestamps to error counts from
                                    that calendar day. Surfaces "deploy day"
                                    correlation.

Slack delivery:
    Uses ``orchestrator/slack_bot.py:post_file`` (the same uploader the
    cost digest uses) plus ``send_dm`` for the narrative line, addressed
    to every user ID in ``SLACK_ADMIN_USER_IDS``. Falls back to direct
    ``slack_sdk.WebClient`` use only if those helpers cannot be imported
    in the runner environment.

Failure handling:
    Failure to DM does not delete the .xlsx; the file path is logged so
    the operator can retrieve it manually. Failure to connect to Postgres
    is loud and returns exit code 2.

Runbook:
    docs/runbooks/README.md  (operator workflow — interpret the
    "Deploys vs incidents" sheet. If a workday error-rate cluster
    appears, that's the data to re-introduce a business-hours deploy
    freeze; see Plan #42 v1 for the design.)

Exit codes:
    0  — success (xlsx written, DM sent or skipped per --dry-run).
    1  — internal error (caught exception; details in stderr).
    2  — Postgres unreachable.
    3  — no session_costs rows in the window (still writes an empty xlsx
         and exits 3 so the cron job is visibly degraded, not silently
         passing on a broken pipeline).
"""

from __future__ import annotations

import argparse
import logging
import os
import subprocess
import sys
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
ORCH_DIR = REPO_ROOT / "orchestrator"
sys.path.insert(0, str(ORCH_DIR))


log = logging.getLogger("measure_deploy_risk")


# ───────────────────────────────────────────────────────────────────────
# Setup helpers
# ───────────────────────────────────────────────────────────────────────


def _load_env() -> None:
    """Manual dotenv loader matching ``orchestrator/config.py``.

    Tests can inject env vars before invoking main; setdefault preserves
    shell-injected values.
    """
    dotenv = REPO_ROOT / ".env"
    if not dotenv.exists():
        return
    for line in dotenv.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip())


def _setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )


# ───────────────────────────────────────────────────────────────────────
# Data access
# ───────────────────────────────────────────────────────────────────────


def _connect_pg():
    """Open a Postgres connection from DATABASE_URL.

    Imported lazily so unit tests can monkey-patch ``psycopg2.connect``
    without paying the import cost up front.
    """
    import psycopg2  # noqa: WPS433 — local import is intentional

    db_url = os.environ.get("DATABASE_URL", "")
    if not db_url:
        raise RuntimeError(
            "DATABASE_URL is unset. Set it locally via .env or on Railway."
        )
    return psycopg2.connect(db_url)


def _fetch_sessions(conn, *, days: int, portco_key: str) -> list[dict]:
    """Return one row per session_costs row in the window.

    Filters to ``portco_key`` so multi-portco installs measure the right
    customer. Returns columns the histogram needs and nothing else —
    keeps memory flat for very wide windows.
    """
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    sql = """
        SELECT
            session_id,
            recorded_at,
            outcome,
            trigger,
            cost_usd
        FROM session_costs
        WHERE recorded_at >= %s
          AND portco_key = %s
        ORDER BY recorded_at ASC
    """
    with conn.cursor() as cur:
        cur.execute(sql, (cutoff, portco_key))
        rows = cur.fetchall()
    return [
        {
            "session_id": r[0],
            "recorded_at": r[1],
            "outcome": r[2] or "success",
            "trigger": r[3] or "",
            "cost_usd": float(r[4] or 0.0),
        }
        for r in rows
    ]


def _fetch_main_commits(days: int) -> list[dict]:
    """Run ``git log`` for main commits in the window.

    Returns ``[{commit, isodate, subject}]``. Quiet on git failures
    (e.g. CI container without a full clone) — returns an empty list.
    """
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    try:
        out = subprocess.check_output(
            [
                "git",
                "-C",
                str(REPO_ROOT),
                "log",
                f"--since={cutoff}",
                "--pretty=format:%H|%cI|%s",
                "main",
            ],
            text=True,
            stderr=subprocess.DEVNULL,
        )
    except (subprocess.CalledProcessError, FileNotFoundError) as exc:
        log.warning(
            "git log failed (%s) — Deploys vs incidents sheet will be empty", exc
        )
        return []
    rows = []
    for line in out.splitlines():
        if not line.strip():
            continue
        try:
            commit, isodate, subject = line.split("|", 2)
        except ValueError:
            continue
        rows.append({"commit": commit, "isodate": isodate, "subject": subject})
    return rows


# ───────────────────────────────────────────────────────────────────────
# Aggregation
# ───────────────────────────────────────────────────────────────────────


def _hourly_session_counts(rows: list[dict]) -> list[tuple[int, int]]:
    """Returns 24 rows: (hour_of_day_PT, session_count)."""
    # Convert UTC timestamps to America/Los_Angeles for display. Avoids a
    # zoneinfo import on Python 3.9 — use a fixed -7/-8 offset is wrong
    # for DST, so we use zoneinfo with a fallback to UTC.
    bins = [0] * 24
    tz = _pacific_tz()
    for r in rows:
        ts = r["recorded_at"]
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        local = ts.astimezone(tz) if tz else ts
        bins[local.hour] += 1
    return [(h, bins[h]) for h in range(24)]


def _hourly_error_rate(rows: list[dict]) -> list[tuple[int, int, int, float]]:
    """Returns 24 rows: (hour, total, errors, error_rate_pct)."""
    total = [0] * 24
    errors = [0] * 24
    tz = _pacific_tz()
    for r in rows:
        ts = r["recorded_at"]
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        local = ts.astimezone(tz) if tz else ts
        h = local.hour
        total[h] += 1
        if r["outcome"] == "error":
            errors[h] += 1
    out = []
    for h in range(24):
        rate = (100.0 * errors[h] / total[h]) if total[h] else 0.0
        out.append((h, total[h], errors[h], round(rate, 2)))
    return out


def _deploys_vs_incidents(
    sessions: list[dict], commits: list[dict]
) -> list[tuple[str, int, int, int]]:
    """Per calendar day: (date, deploy_count, session_count, error_count)."""
    tz = _pacific_tz()
    sessions_by_day: dict[str, list[dict]] = {}
    for r in sessions:
        ts = r["recorded_at"]
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        local = ts.astimezone(tz) if tz else ts
        key = local.strftime("%Y-%m-%d")
        sessions_by_day.setdefault(key, []).append(r)

    deploys_by_day: dict[str, int] = {}
    for c in commits:
        try:
            dt = datetime.fromisoformat(c["isodate"])
        except ValueError:
            continue
        local = dt.astimezone(tz) if tz else dt
        key = local.strftime("%Y-%m-%d")
        deploys_by_day[key] = deploys_by_day.get(key, 0) + 1

    all_days = sorted(set(sessions_by_day) | set(deploys_by_day))
    out = []
    for day in all_days:
        ds = sessions_by_day.get(day, [])
        errors = sum(1 for r in ds if r["outcome"] == "error")
        out.append((day, deploys_by_day.get(day, 0), len(ds), errors))
    return out


def _pacific_tz():
    """Return America/Los_Angeles tzinfo when available, else UTC.

    Python 3.9 ships ``zoneinfo`` only when ``tzdata`` is present;
    in CI we may need to fall back. We don't fail the whole script
    over a timezone lookup — the histogram is still useful in UTC.
    """
    try:
        from zoneinfo import ZoneInfo  # noqa: WPS433

        return ZoneInfo("America/Los_Angeles")
    except Exception:  # pragma: no cover — depends on runner OS
        log.warning("zoneinfo unavailable — falling back to UTC for hour bins")
        return None


# ───────────────────────────────────────────────────────────────────────
# Workbook output
# ───────────────────────────────────────────────────────────────────────


def _write_xlsx(
    out_path: Path,
    *,
    sessions: list[tuple[int, int]],
    error_rate: list[tuple[int, int, int, float]],
    deploys: list[tuple[str, int, int, int]],
    portco_key: str,
    days: int,
) -> None:
    """Emit the 3-sheet workbook with a metadata header on each tab."""
    from openpyxl import Workbook  # noqa: WPS433

    wb = Workbook()
    # Default sheet → Sessions by hour
    ws1 = wb.active
    ws1.title = "Sessions by hour"
    ws1.append(
        [
            f"Portco: {portco_key}",
            f"Window: last {days} days",
            "tz: America/Los_Angeles",
        ]
    )
    ws1.append([])
    ws1.append(["Hour (PT)", "Session count"])
    for h, n in sessions:
        ws1.append([f"{h:02d}", n])

    ws2 = wb.create_sheet("Error rate by hour")
    ws2.append([f"Portco: {portco_key}", f"Window: last {days} days"])
    ws2.append([])
    ws2.append(["Hour (PT)", "Sessions", "Errors", "Error rate (%)"])
    for h, total, errors, rate in error_rate:
        ws2.append([f"{h:02d}", total, errors, rate])

    ws3 = wb.create_sheet("Deploys vs incidents")
    ws3.append([f"Portco: {portco_key}", f"Window: last {days} days"])
    ws3.append([])
    ws3.append(["Date (PT)", "Main commits", "Sessions", "Errors"])
    for day, deploys_n, sessions_n, errors_n in deploys:
        ws3.append([day, deploys_n, sessions_n, errors_n])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


# ───────────────────────────────────────────────────────────────────────
# Slack delivery
# ───────────────────────────────────────────────────────────────────────


def _send_admin_dm_with_file(
    out_path: Path,
    *,
    summary: str,
    dry_run: bool,
) -> int:
    """DM every admin user the .xlsx + a one-line summary.

    Returns the number of admins successfully messaged. Never raises —
    the calling code already wrote the file, so a Slack failure should
    not change the script's exit code.
    """
    if dry_run:
        log.info(
            "[DRY-RUN] Would DM admins with file=%s and summary=%s", out_path, summary
        )
        return 0

    try:
        from cost_digest import _resolve_admin_ids  # noqa: WPS433
    except Exception as exc:
        log.warning("Could not import _resolve_admin_ids: %s", exc)
        return 0

    try:
        admins = _resolve_admin_ids()
    except Exception as exc:
        log.warning("Failed to resolve admin IDs: %s", exc)
        return 0
    if not admins:
        log.warning(
            "No admins configured (SLACK_ADMIN_USER_IDS unset) — file at %s", out_path
        )
        return 0

    # Use slack_sdk directly so we control the per-user DM upload exactly.
    # post_file in slack_bot defaults to the public channel; admin DM
    # delivery needs per-user channel resolution.
    try:
        from slack_sdk import WebClient  # noqa: WPS433
    except Exception as exc:
        log.warning("slack_sdk import failed (%s) — file kept at %s", exc, out_path)
        return 0

    token = os.environ.get("SLACK_BOT_TOKEN", "")
    if not token:
        log.warning("SLACK_BOT_TOKEN unset — file kept at %s", out_path)
        return 0
    client = WebClient(token=token)

    sent = 0
    for uid in admins:
        try:
            dm = client.conversations_open(users=[uid])
            channel = dm["channel"]["id"]
            client.files_upload_v2(
                channel=channel,
                file=str(out_path),
                filename=out_path.name,
                title=out_path.stem,
                initial_comment=summary,
            )
            sent += 1
            log.info("DM-OK sent to %s", uid)
        except Exception as exc:
            log.warning("DM-FAIL %s: %s", uid, exc)
    return sent


# ───────────────────────────────────────────────────────────────────────
# Main
# ───────────────────────────────────────────────────────────────────────


def _build_summary(
    *,
    portco_key: str,
    days: int,
    total_sessions: int,
    total_errors: int,
    busiest_hour: int,
    rate_at_busiest: float,
) -> str:
    """Compose the admin DM body. Pattern: WHAT + WHY + FIX-COMMAND."""
    error_rate = (100.0 * total_errors / total_sessions) if total_sessions else 0.0
    return (
        f"[DEPLOY-RISK MEASUREMENT] {portco_key} — last {days} days\n"
        f"  Sessions:        {total_sessions:,}\n"
        f"  Errors:          {total_errors:,} ({error_rate:.1f}%)\n"
        f"  Busiest hour PT: {busiest_hour:02d} ({rate_at_busiest:.1f}% error rate)\n\n"
        f"Workbook attached. Read the 'Error rate by hour' tab first —\n"
        f"if errors cluster during Acme's workday (09:00–18:00 ET =\n"
        f"06:00–15:00 PT), that's the data point to re-introduce a\n"
        f"business-hours deploy freeze (Plan #42 v1 design).\n\n"
        f"Runbook: docs/runbooks/README.md\n"
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="measure-deploy-risk",
        description=(
            "Measure deploy risk for the Acme GTM Health Agent. "
            "Reads session_costs for the last N days and emits an .xlsx "
            "histogram + admin Slack DM. See docs/runbooks/README.md "
            "for how to interpret the output."
        ),
        epilog=(
            "Examples:\n"
            "  bin/measure-deploy-risk.py\n"
            "  bin/measure-deploy-risk.py --days 30 --out /tmp/foo.xlsx\n"
            "  bin/measure-deploy-risk.py --portco acme --dry-run\n"
            "\n"
            "Runbook: docs/runbooks/README.md"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--days",
        type=int,
        default=60,
        help="Number of days back to include (default: 60).",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help=(
            "Output .xlsx path (default: /tmp/deploy_risk_<DATE>.xlsx in your tmp dir)."
        ),
    )
    parser.add_argument(
        "--portco",
        default="acme",
        help="Portco key to measure (default: acme).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Compute and write the .xlsx but do not DM admins.",
    )
    args = parser.parse_args(argv)

    _setup_logging()
    _load_env()

    if args.out is None:
        date_str = datetime.now().strftime("%Y%m%d")
        args.out = Path("/tmp") / f"deploy_risk_{date_str}.xlsx"

    log.info(
        "Measuring deploy risk: portco=%s days=%d out=%s",
        args.portco,
        args.days,
        args.out,
    )

    try:
        conn = _connect_pg()
    except Exception as exc:
        log.error("Postgres unreachable: %s", exc)
        return 2

    try:
        sessions = _fetch_sessions(conn, days=args.days, portco_key=args.portco)
    finally:
        try:
            conn.close()
        except Exception:
            pass

    commits = _fetch_main_commits(args.days)

    sessions_per_hour = _hourly_session_counts(sessions)
    error_rate_per_hour = _hourly_error_rate(sessions)
    deploys = _deploys_vs_incidents(sessions, commits)

    _write_xlsx(
        args.out,
        sessions=sessions_per_hour,
        error_rate=error_rate_per_hour,
        deploys=deploys,
        portco_key=args.portco,
        days=args.days,
    )
    log.info("Wrote %s", args.out)

    total_sessions = sum(n for _, n in sessions_per_hour)
    total_errors = sum(e for _, _, e, _ in error_rate_per_hour)
    # Busiest hour by sessions, then by error rate as tiebreaker.
    if total_sessions == 0:
        busiest_hour, rate_at_busiest = 0, 0.0
    else:
        busiest = max(
            error_rate_per_hour,
            key=lambda row: (row[1], row[3]),
        )
        busiest_hour, rate_at_busiest = busiest[0], busiest[3]

    summary = _build_summary(
        portco_key=args.portco,
        days=args.days,
        total_sessions=total_sessions,
        total_errors=total_errors,
        busiest_hour=busiest_hour,
        rate_at_busiest=rate_at_busiest,
    )

    sent = _send_admin_dm_with_file(args.out, summary=summary, dry_run=args.dry_run)
    log.info("Admin DM delivery: %d sent", sent)

    # Operator-facing "Next steps" block — points back at the runbook
    # so the cron output is actionable even without reading the source.
    print()
    print("Next steps:")
    print(f"  1. Open {args.out} (.xlsx, 3 sheets).")
    print("  2. Read the 'Error rate by hour' tab first.")
    print("  3. If the hottest hour falls outside 06:00 – 15:00 PT, propose")
    print(
        "     re-introducing a business-hours freeze if errors cluster during the workday (Plan #42 v1 design)."
    )
    print("  4. Runbook: docs/runbooks/README.md")
    print()

    if total_sessions == 0:
        log.warning(
            "No session_costs rows in the window — exit 3 to signal degraded data"
        )
        return 3
    return 0


if __name__ == "__main__":  # pragma: no cover
    try:
        sys.exit(main())
    except SystemExit:
        raise
    except Exception:
        traceback.print_exc()
        sys.exit(1)
