"""Tests for ``bin/measure-deploy-risk.py`` (Plan #42 PR1).

Fully mocked — no real Postgres, no real Slack. The script has a hyphen
in its filename, so we load it by path via importlib.util (same pattern
as ``bin/rollback_agent_test.py``).

Covered:
- ``--help`` succeeds and references the runbook.
- ``--dry-run`` writes a 3-sheet .xlsx and does NOT call Slack.
- An admin user with a valid token receives a DM with the file attached.
- Hourly aggregation buckets rows by Pacific hour-of-day.
- Empty result set returns exit code 3.
- Postgres unavailable returns exit code 2.
"""

from __future__ import annotations

import importlib.util
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPT_PATH = REPO_ROOT / "bin" / "measure-deploy-risk.py"


# Same env bootstrap as bin/backfill_lead_sync_fields_test.py — config.py
# raises on missing keys at module load.
for _key, _value in {
    "ANTHROPIC_API_KEY": "sk-ant-test",
    "SLACK_BOT_TOKEN": "xoxb-test",
    "SLACK_APP_TOKEN": "xapp-test",
    "SLACK_CHANNEL_ID": "C0TEST",
    "ENVIRONMENT_ID": "env_test",
    "DREAM_AGENT_ID": "agent_test_dream",
    "COORDINATOR_ID": "agent_test_coord",
    "QUICK_AGENT_ID": "agent_test_quick",
    "METHODOLOGY_STORE_ID": "memstore_test_m",
    "HEALTH_STORE_ID": "memstore_test_h",
    "DATABASE_URL": "postgres://test/db",
    "SLACK_ADMIN_USER_IDS": "U_ADMIN_1",
}.items():
    os.environ.setdefault(_key, _value)


def _load_script_module():
    """Load ``bin/measure-deploy-risk.py`` by path."""
    for p in (REPO_ROOT / "orchestrator", REPO_ROOT / "bin"):
        if str(p) not in sys.path:
            sys.path.insert(0, str(p))
    spec = importlib.util.spec_from_file_location("measure_deploy_risk", SCRIPT_PATH)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture()
def mdr_mod():
    return _load_script_module()


def _ts(hours_ago: int = 0, days_ago: int = 0):
    """Return a UTC datetime offset by N hours/days."""
    return datetime.now(timezone.utc) - timedelta(hours=hours_ago, days=days_ago)


def _mock_pg_with_rows(rows):
    """Build a psycopg2-shaped MagicMock with the given rows.

    ``rows`` is a list of tuples matching the SELECT in ``_fetch_sessions``:
    (session_id, recorded_at, outcome, trigger, cost_usd).
    """
    cur = MagicMock()
    cur.fetchall.return_value = rows
    conn = MagicMock()
    conn.cursor.return_value.__enter__.return_value = cur
    return conn


# ───────────────────────────────────────────────────────────────────────
# --help
# ───────────────────────────────────────────────────────────────────────


def test_help_succeeds_and_mentions_runbook(mdr_mod, capsys):
    """``--help`` must exit 0 and reference the business-hours runbook
    so an operator running ``bin/measure-deploy-risk.py --help`` at 2 AM
    is told where to read more.
    """
    with pytest.raises(SystemExit) as exc:
        mdr_mod.main(["--help"])
    assert exc.value.code == 0
    out = capsys.readouterr().out
    assert "runbooks/README.md" in out


# ───────────────────────────────────────────────────────────────────────
# Aggregation helpers
# ───────────────────────────────────────────────────────────────────────


def test_hourly_session_counts_buckets_by_hour(mdr_mod):
    """Rows at 09:00 PT and 10:00 PT land in their respective hour bins."""
    tz = mdr_mod._pacific_tz()
    if tz is None:
        pytest.skip("zoneinfo unavailable on this runner")
    nine_pt = datetime(2026, 5, 1, 9, 0, tzinfo=tz)
    ten_pt = datetime(2026, 5, 1, 10, 0, tzinfo=tz)
    rows = [
        {
            "recorded_at": nine_pt.astimezone(timezone.utc),
            "outcome": "success",
            "trigger": "",
            "cost_usd": 0.0,
            "session_id": "a",
        },
        {
            "recorded_at": nine_pt.astimezone(timezone.utc),
            "outcome": "error",
            "trigger": "",
            "cost_usd": 0.0,
            "session_id": "b",
        },
        {
            "recorded_at": ten_pt.astimezone(timezone.utc),
            "outcome": "success",
            "trigger": "",
            "cost_usd": 0.0,
            "session_id": "c",
        },
    ]
    bins = mdr_mod._hourly_session_counts(rows)
    by_hour = dict(bins)
    assert by_hour[9] == 2
    assert by_hour[10] == 1


def test_hourly_error_rate_computes_pct(mdr_mod):
    """Half-success / half-error → 50% error rate."""
    tz = mdr_mod._pacific_tz()
    if tz is None:
        pytest.skip("zoneinfo unavailable on this runner")
    nine_pt = datetime(2026, 5, 1, 9, 0, tzinfo=tz)
    rows = [
        {
            "recorded_at": nine_pt.astimezone(timezone.utc),
            "outcome": "success",
            "trigger": "",
            "cost_usd": 0.0,
            "session_id": "a",
        },
        {
            "recorded_at": nine_pt.astimezone(timezone.utc),
            "outcome": "error",
            "trigger": "",
            "cost_usd": 0.0,
            "session_id": "b",
        },
    ]
    out = mdr_mod._hourly_error_rate(rows)
    by_hour = {h: (t, e, r) for h, t, e, r in out}
    total, errors, rate = by_hour[9]
    assert total == 2 and errors == 1 and rate == 50.0


# ───────────────────────────────────────────────────────────────────────
# main() end-to-end with mocks
# ───────────────────────────────────────────────────────────────────────


def test_main_dry_run_writes_xlsx_and_skips_slack(mdr_mod, tmp_path):
    """``--dry-run`` writes the workbook but does NOT call Slack."""
    out_path = tmp_path / "test_dry.xlsx"
    rows = [
        ("sess-1", _ts(hours_ago=1), "success", "slack", 0.01),
        ("sess-2", _ts(hours_ago=2), "error", "slack", 0.02),
        ("sess-3", _ts(hours_ago=3), "success", "cron", 0.01),
    ]
    with (
        patch.object(mdr_mod, "_connect_pg", return_value=_mock_pg_with_rows(rows)),
        patch.object(mdr_mod, "_fetch_main_commits", return_value=[]),
    ):
        slack_calls = []

        def fake_dm(*args, **kwargs):
            slack_calls.append((args, kwargs))
            return 0

        with patch.object(
            mdr_mod, "_send_admin_dm_with_file", side_effect=fake_dm
        ) as mock_dm:
            code = mdr_mod.main(
                [
                    "--days",
                    "60",
                    "--out",
                    str(out_path),
                    "--dry-run",
                ]
            )
        # _send_admin_dm_with_file IS called (it's where the dry-run check
        # short-circuits) but with dry_run=True so no Slack calls fire.
        assert mock_dm.called
        kwargs = mock_dm.call_args.kwargs
        assert kwargs["dry_run"] is True

    assert code == 0
    assert out_path.exists() and out_path.stat().st_size > 0

    # Sanity-check the workbook structure: 3 sheets in the documented order.
    from openpyxl import load_workbook

    wb = load_workbook(str(out_path))
    assert wb.sheetnames == [
        "Sessions by hour",
        "Error rate by hour",
        "Deploys vs incidents",
    ]


def test_main_calls_slack_dm_with_xlsx_when_admin_configured(mdr_mod, tmp_path):
    """End-to-end with admin configured: WebClient.files_upload_v2 is invoked
    on the per-admin DM channel with the .xlsx attached.
    """
    out_path = tmp_path / "test_dm.xlsx"
    rows = [
        ("sess-1", _ts(hours_ago=1), "success", "slack", 0.01),
        ("sess-2", _ts(hours_ago=2), "error", "slack", 0.02),
    ]

    # Mock the slack_sdk WebClient + cost_digest admin resolver.
    fake_client = MagicMock()
    fake_client.conversations_open.return_value = {"channel": {"id": "D123"}}
    fake_client.files_upload_v2.return_value = {"ok": True}

    fake_sdk = SimpleNamespace(WebClient=MagicMock(return_value=fake_client))
    fake_cost_digest = SimpleNamespace(_resolve_admin_ids=lambda: ["U_ADMIN_1"])

    with (
        patch.object(mdr_mod, "_connect_pg", return_value=_mock_pg_with_rows(rows)),
        patch.object(mdr_mod, "_fetch_main_commits", return_value=[]),
        patch.dict(
            sys.modules, {"slack_sdk": fake_sdk, "cost_digest": fake_cost_digest}
        ),
    ):
        code = mdr_mod.main(
            [
                "--days",
                "30",
                "--out",
                str(out_path),
                "--portco",
                "acme",
            ]
        )

    assert code == 0
    assert out_path.exists()
    fake_client.files_upload_v2.assert_called_once()
    upload_kwargs = fake_client.files_upload_v2.call_args.kwargs
    assert upload_kwargs["channel"] == "D123"
    assert upload_kwargs["file"] == str(out_path)


def test_main_returns_3_when_no_sessions(mdr_mod, tmp_path):
    """Empty window → exit code 3 (visibly degraded, not silent pass)."""
    out_path = tmp_path / "test_empty.xlsx"
    with (
        patch.object(mdr_mod, "_connect_pg", return_value=_mock_pg_with_rows([])),
        patch.object(mdr_mod, "_fetch_main_commits", return_value=[]),
        patch.object(mdr_mod, "_send_admin_dm_with_file", return_value=0),
    ):
        code = mdr_mod.main(
            [
                "--out",
                str(out_path),
                "--dry-run",
            ]
        )
    assert code == 3
    assert out_path.exists()  # still wrote the .xlsx so operators can inspect


def test_main_returns_2_when_pg_unreachable(mdr_mod, tmp_path):
    """Postgres connect failure → exit code 2."""
    out_path = tmp_path / "test_pg_fail.xlsx"

    def _raise(*args, **kwargs):
        raise RuntimeError("DATABASE_URL is unset.")

    with patch.object(mdr_mod, "_connect_pg", side_effect=_raise):
        code = mdr_mod.main(
            [
                "--out",
                str(out_path),
                "--dry-run",
            ]
        )
    assert code == 2


def test_summary_includes_runbook_pointer(mdr_mod):
    """The DM body must reference the runbook so admins see the link."""
    body = mdr_mod._build_summary(
        portco_key="acme",
        days=60,
        total_sessions=100,
        total_errors=5,
        busiest_hour=10,
        rate_at_busiest=12.5,
    )
    assert "runbooks/README.md" in body
    assert "100" in body
    assert "5" in body
