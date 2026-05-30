"""Tests for orchestrator/xlsx_export.parquet_to_xlsx_sibling.

Covers the happy path (Parquet → xlsx round-trip preserves cells + column
order), the streaming case (50K-row Parquet doesn't blow up RSS), and the
silent-fallback contract (any failure returns None, never raises, leaves
Parquet path intact).
"""

from __future__ import annotations

import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq

ORCH = Path(__file__).resolve().parent
if str(ORCH) not in sys.path:
    sys.path.insert(0, str(ORCH))


def _write_parquet(path: str, table: pa.Table) -> None:
    """Write a pyarrow Table to ``path`` as Parquet."""
    pq.write_table(table, path)


def test_parquet_to_xlsx_writes_sibling_with_matching_rows(tmp_path):
    """Happy path: Parquet on disk → xlsx with same rows, same column order."""
    from xlsx_export import parquet_to_xlsx_sibling

    table = pa.table(
        {
            "Id": ["00Q1", "00Q2", "00Q3"],
            "Name": ["Alice", "Bob", "Carol"],
            "Status": ["Working", "Disqualified", "Working"],
        }
    )
    parquet_path = str(tmp_path / "leads_test.parquet")
    _write_parquet(parquet_path, table)

    xlsx_path = parquet_to_xlsx_sibling(parquet_path)
    assert xlsx_path is not None
    assert xlsx_path == str(tmp_path / "leads_test.xlsx")
    assert os.path.exists(xlsx_path)

    # Read back via openpyxl and verify cells + column order match.
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["data"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Id", "Name", "Status"), "header row preserves column order"
    assert rows[1] == ("00Q1", "Alice", "Working")
    assert rows[2] == ("00Q2", "Bob", "Disqualified")
    assert rows[3] == ("00Q3", "Carol", "Working")
    wb.close()


def test_parquet_to_xlsx_coerces_pyarrow_timestamps_to_python_datetime(tmp_path):
    """pyarrow Timestamp cells must become Python datetimes — openpyxl rejects pyarrow types."""
    from xlsx_export import parquet_to_xlsx_sibling

    dt1 = datetime(2025, 1, 15, 10, 30, 0, tzinfo=timezone.utc)
    dt2 = datetime(2025, 6, 1, 0, 0, 0, tzinfo=timezone.utc)
    table = pa.table(
        {
            "Id": ["A", "B"],
            "CreatedDate": pa.array([dt1, dt2], type=pa.timestamp("us", tz="UTC")),
        }
    )
    parquet_path = str(tmp_path / "dates.parquet")
    _write_parquet(parquet_path, table)

    xlsx_path = parquet_to_xlsx_sibling(parquet_path)
    assert xlsx_path is not None

    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["data"]
    rows = list(ws.iter_rows(values_only=True))
    # openpyxl reads back datetime cells as datetime objects.
    assert rows[0] == ("Id", "CreatedDate")
    assert isinstance(rows[1][1], datetime)
    assert rows[1][1].year == 2025 and rows[1][1].month == 1
    assert isinstance(rows[2][1], datetime)
    wb.close()


def test_parquet_to_xlsx_returns_none_for_missing_file(tmp_path):
    """A path that doesn't exist returns None, never raises."""
    from xlsx_export import parquet_to_xlsx_sibling

    result = parquet_to_xlsx_sibling(str(tmp_path / "does_not_exist.parquet"))
    assert result is None


def test_parquet_to_xlsx_returns_none_for_non_parquet_path(tmp_path):
    """A path that isn't .parquet returns None — no work to do."""
    from xlsx_export import parquet_to_xlsx_sibling

    fake = tmp_path / "data.csv"
    fake.write_text("Id,Name\n1,Alice\n")
    assert parquet_to_xlsx_sibling(str(fake)) is None


def test_parquet_to_xlsx_returns_none_for_empty_or_none_path():
    """Empty / None inputs return None — defensive against caller bugs."""
    from xlsx_export import parquet_to_xlsx_sibling

    assert parquet_to_xlsx_sibling("") is None
    assert parquet_to_xlsx_sibling(None) is None  # type: ignore[arg-type]


def test_parquet_to_xlsx_streams_large_parquet_without_loading_all_rows(tmp_path):
    """50K-row Parquet writes successfully via batched iteration.

    The batched read path is the whole point of write_only mode — pulling
    everything into memory at once breaks Railway's 512 MB container cap
    on multi-hundred-thousand-row dumps. This test forces the streaming
    path by exceeding the default batch size and verifies the round-trip.
    """
    from xlsx_export import parquet_to_xlsx_sibling

    n = 50_000
    table = pa.table(
        {
            "Id": [f"00Q{i:05d}" for i in range(n)],
            "Value": list(range(n)),
        }
    )
    parquet_path = str(tmp_path / "big.parquet")
    _write_parquet(parquet_path, table)

    xlsx_path = parquet_to_xlsx_sibling(parquet_path)
    assert xlsx_path is not None
    assert os.path.exists(xlsx_path)

    # Spot-check first + last rows; loading all 50K into Python would be slow.
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["data"]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    assert header == ("Id", "Value")
    first = next(rows_iter)
    assert first == ("00Q00000", 0)
    # Walk to the last row.
    last = first
    for row in rows_iter:
        last = row
    assert last == (f"00Q{n - 1:05d}", n - 1)
    wb.close()


def test_parquet_to_xlsx_cleans_up_partial_file_on_write_failure(monkeypatch, tmp_path):
    """If wb.save raises, no partial xlsx is left on disk.

    Half-written files would confuse the attach-side sibling lookup
    (``_prefer_xlsx_sibling`` swaps to xlsx if it exists, so a corrupt
    xlsx would be uploaded instead of the Parquet). Best-effort cleanup
    is part of the silent-fallback contract.
    """
    from xlsx_export import parquet_to_xlsx_sibling

    table = pa.table({"Id": ["A"], "Name": ["x"]})
    parquet_path = str(tmp_path / "willfail.parquet")
    _write_parquet(parquet_path, table)

    # Force a save failure mid-write by replacing Workbook.save with a raise.
    import openpyxl

    original_save = openpyxl.Workbook.save

    def _broken_save(self, filename):
        # Simulate a partial write — touch the file first, then raise.
        with open(filename, "wb") as f:
            f.write(b"partial")
        raise IOError("disk full")

    monkeypatch.setattr(openpyxl.Workbook, "save", _broken_save)
    try:
        result = parquet_to_xlsx_sibling(parquet_path)
        assert result is None
        # Cleanup must have removed the partial file.
        partial = str(tmp_path / "willfail.xlsx")
        assert not os.path.exists(partial), (
            "partial xlsx must be cleaned up so the attach-side sibling lookup "
            "doesn't pick up a corrupt file"
        )
    finally:
        monkeypatch.setattr(openpyxl.Workbook, "save", original_save)
