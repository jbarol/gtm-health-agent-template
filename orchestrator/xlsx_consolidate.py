"""Merge multiple ``.xlsx`` files into one workbook before Slack upload.

Plan #48 Phase 1 — module only. The wiring into ``_dispatch_post_report``
and the split-files registry land in later phases. This module is a pure,
testable transform from N xlsx paths to either one consolidated xlsx path
or the original N paths (size-cap fallback or any failure).

Design contract:

  * **Never raises.** Every failure mode returns ``(original_paths, False)``
    and logs a structured line. The consolidation is best-effort UX
    polish; falling back to N individual files is safe.

  * **Streaming I/O on both sides.** Read with
    ``load_workbook(path, read_only=True)``, write with
    ``Workbook(write_only=True)``. The combined memory peak is roughly
    one row batch of the widest sheet, not the total across all sheets.

  * **Banner row on every sheet.** Row 1 is ``Source: <basename>`` so an
    operator opening the consolidated file knows where each sheet came
    from. Row 2 is column headers (copied from source row 1). Row 3+ is
    data (copied from source row 2+).

  * **Sheet name derivation strips machine-generated noise.** SF dump
    siblings look like ``sf_pipeline_opps_2026_<20-digit-ts>_<4-hex>.xlsx``;
    the user-facing sheet should read ``Pipeline Opps 2026``, not the raw
    stem. Coordinator-named files (e.g. ``call_prep_brief.xlsx``) keep
    their stem.

  * **Size cap protects the 512 MB Railway container.** Sum input sizes
    before opening any file. Above the cap (default 50 MB), log
    ``[XLSX_CONSOLIDATE_SKIPPED_SIZE]`` and return original paths.

  * **Output is overwrite-safe.** A ``post_report`` retry rebuilds the
    attachment list and re-calls this function; the consolidated file
    name is deterministic by date so the retry overwrites cleanly.
"""

from __future__ import annotations

import datetime as _dt
import logging
import os
import re

log = logging.getLogger(__name__)


# Machine-generated SF dump filename pattern (from sf_dump_tool._utc_iso_compact:
# strftime("%Y%m%dT%H%M%S%f") → 8 + 1 + 6 + 6 = 21 chars; suffix is
# uuid4().hex[:4]). The label segment is non-greedy so the regex anchors on
# the timestamp + hex tail rather than over-eating into the label.
_SF_DUMP_FILENAME_RE = re.compile(
    r"^sf_(?P<label>.+?)_\d{8}T\d{12}_[0-9a-f]{4}\.xlsx$",
    re.IGNORECASE,
)

# Excel-forbidden sheet name chars.
_EXCEL_FORBIDDEN = set(r":\/?*[]")


def consolidate_xlsx_files(
    paths: list[str],
    output_dir: str,
    portco_key: str = "unknown",
    size_cap_mb: float = 50.0,
) -> tuple[list[str], bool]:
    """Merge xlsx files into one workbook.

    Returns ``(merged_path_list, did_consolidate)``. On success, the list
    contains exactly one path; on no-op or failure, the original list is
    returned with ``did_consolidate=False``. Never raises.
    """
    # Empty input: nothing to do, no consolidation claimed.
    if not paths:
        return [], False

    # Single file: no merge needed.
    if len(paths) == 1:
        return list(paths), False

    try:
        # Size cap — guard before opening anything.
        cap_bytes = int(size_cap_mb * 1_048_576)
        total = 0
        for p in paths:
            try:
                total += os.path.getsize(p)
            except OSError:
                # Missing or unreadable; let downstream openpyxl raise.
                pass
        if total > cap_bytes:
            log.warning(
                "[XLSX_CONSOLIDATE_SKIPPED_SIZE] portco=%s total_mb=%.2f cap_mb=%.2f files=%s",
                portco_key,
                total / 1_048_576,
                size_cap_mb,
                [os.path.basename(p) for p in paths],
            )
            return list(paths), False

        from openpyxl import Workbook, load_workbook

        ordered = _sort_for_consolidation(paths)

        out_path = _derive_output_path(output_dir, portco_key)

        wb = Workbook(write_only=True)
        existing: set[str] = set()
        for src in ordered:
            base = _derive_sheet_name(
                src, set()
            )  # candidate base; dedup happens per-sheet below
            src_wb = load_workbook(src, read_only=True)
            try:
                sheets = src_wb.worksheets
                if not sheets:
                    continue
                # Codex P2 fix (2026-05-18): when a source xlsx is itself
                # multi-sheet (e.g. from materialize_xlsx(sheets=[...]) on the
                # Coordinator path), preserve every sheet. Previously only
                # worksheets[0] was copied; subsequent tabs were silently
                # dropped from the consolidated output.
                for src_ws in sheets:
                    if len(sheets) == 1:
                        candidate = base
                    else:
                        # Use the source sheet's own title (sanitized) so the
                        # consolidated workbook keeps the materialize_xlsx
                        # author's intended labels. Fall back to base + index
                        # suffix when the title is empty.
                        title = _sanitize_sheet_name(src_ws.title or "")
                        candidate = title or base
                    sheet_name = _dedup_sheet_name(candidate, existing)
                    existing.add(sheet_name)
                    _copy_sheet(src_ws, wb, sheet_name, os.path.basename(src))
            finally:
                src_wb.close()

        wb.save(out_path)
        log.info(
            "[XLSX_CONSOLIDATED] portco=%s files_merged=%d output=%s",
            portco_key,
            len(ordered),
            out_path,
        )
        return [out_path], True

    except Exception:
        log.exception("[XLSX_CONSOLIDATE_FAILED] portco=%s", portco_key)
        return list(paths), False


def _is_coordinator_produced(path: str) -> bool:
    """True when the filename does NOT match the machine-generated dump pattern.

    Coordinator-named files (e.g. ``call_prep_brief.xlsx``,
    ``pipeline_summary.xlsx``) flow through ``materialize_xlsx`` and keep a
    human-readable stem. SF-dump siblings are always
    ``sf_<label>_<ts>_<hex>.xlsx``. The coordinator-produced ones sort first
    so the user sees the synthesized output before the raw dumps.
    """
    return _SF_DUMP_FILENAME_RE.match(os.path.basename(path)) is None


def _sort_for_consolidation(paths: list[str]) -> list[str]:
    """Coordinator-produced files first, then SF dumps by mtime ascending.

    Tie-break inside each group by mtime so chronological order of
    materialization is preserved.
    """

    def _key(p: str) -> tuple[int, float]:
        coord = 0 if _is_coordinator_produced(p) else 1
        try:
            mtime = os.path.getmtime(p)
        except OSError:
            mtime = 0.0
        return (coord, mtime)

    return sorted(paths, key=_key)


def _sanitize_sheet_name(raw: str) -> str:
    """Replace Excel-forbidden chars with ``_``, strip whitespace, truncate.

    Excel's hard limit is 31 chars; we truncate at 28 to leave room for
    a ``" 2"`` / ``" 3"`` dedup suffix appended by ``_derive_sheet_name``
    on collision.
    """
    if not raw:
        return "Sheet"
    cleaned = "".join("_" if c in _EXCEL_FORBIDDEN else c for c in raw)
    cleaned = cleaned.strip()
    if len(cleaned) > 28:
        cleaned = cleaned[:28].rstrip()
    return cleaned or "Sheet"


def _derive_sheet_name(xlsx_path: str, existing_names: set[str]) -> str:
    """Strip ``sf_`` prefix + ``_<ts>_<hex>`` suffix, title-case, sanitize, dedup."""
    basename = os.path.basename(xlsx_path)
    stem, _ = os.path.splitext(basename)

    m = _SF_DUMP_FILENAME_RE.match(basename)
    if m:
        stem = m.group("label")

    pretty = stem.replace("_", " ").strip()
    pretty = pretty.title() if pretty else "Sheet"

    return _dedup_sheet_name(_sanitize_sheet_name(pretty), existing_names)


def _dedup_sheet_name(base: str, existing_names: set[str]) -> str:
    """Append `" 2"`, `" 3"`, ... when ``base`` collides. ≤31 chars total."""
    if base not in existing_names:
        return base
    n = 2
    while True:
        suffix = f" {n}"
        head_room = 31 - len(suffix)
        candidate = f"{base[:head_room].rstrip()}{suffix}"
        if candidate not in existing_names:
            return candidate
        n += 1


def _copy_sheet(source_ws, dest_wb, sheet_name: str, source_filename: str) -> None:
    """Copy rows from ``source_ws`` into a new write-only sheet on ``dest_wb``.

    Layout:
      row 1: ``Source: <source_filename>`` banner (single cell)
      row 2: column headers (from source row 1)
      row 3+: data (from source rows 2+)

    Read with ``iter_rows(values_only=True)`` so we never materialize the
    whole sheet in memory.
    """
    ws = dest_wb.create_sheet(title=sheet_name)
    ws.append([f"Source: {source_filename}"])

    rows_iter = source_ws.iter_rows(values_only=True)
    for row in rows_iter:
        # openpyxl tuples → lists for write_only append. None cells pass
        # through unchanged.
        ws.append(list(row))


def _derive_output_path(output_dir: str, portco_key: str) -> str:
    """Build a per-call unique consolidated-file path.

    Codex P2 fix (2026-05-18): a deterministic ``<portco>_<date>.xlsx`` name
    collided when two investigations for the same portco ran on the same
    UTC date — the later save overwrote the earlier workbook, and async
    Slack uploads of the earlier file could then read the wrong content.
    The microsecond timestamp + 4-hex random suffix gives ~16M*(1us) call
    capacity per portco-second without collision. The 14-day artifact sweep
    in ``artifact_paths`` cleans these up; retries within the same call
    produce a new file (cheap; the prior orphan is reclaimed).
    """
    import secrets

    stamp = _dt.datetime.now(_dt.timezone.utc).strftime("%Y%m%dT%H%M%S%f")
    rand = secrets.token_hex(2)
    name = f"consolidated_{portco_key}_{stamp}_{rand}.xlsx"
    return os.path.join(output_dir, name)
