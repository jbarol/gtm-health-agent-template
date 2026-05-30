"""Tests for ``orchestrator.result_virtualize``.

Covers:
    - 100-row round trip: dict list → .xlsx → identical readback
    - Summary stats: numeric column min/max/mean/count; low-cardinality string
      column top_values; high-cardinality string column unique_count
    - Zero-row safety: empty input returns the full shape, writes no file
    - 10K-row performance: completes in <5 seconds with no memory blow-up

Run:
    cd orchestrator && python3 -m pytest result_virtualize_test.py -q
"""

from __future__ import annotations

import os
import time
from pathlib import Path

from result_virtualize import (
    LOW_CARDINALITY_MAX,
    PREVIEW_ROW_COUNT,
    compute_summary_stats,
    virtualize_result,
    write_xlsx_streaming,
)


# ──────────────────────────────────────────────────────────────────────────
# write_xlsx_streaming — round-trip
# ──────────────────────────────────────────────────────────────────────────


def _read_xlsx(path: str) -> list[dict]:
    """Read an .xlsx back into a list of row dicts. Reverses write_xlsx_streaming."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    out = []
    for row in rows_iter:
        out.append({h: v for h, v in zip(headers, row)})
    return out


def test_write_xlsx_round_trip_preserves_rows(tmp_path: Path):
    """100 rows written then read back match exactly. Header order preserved."""
    rows = [
        {
            "Id": f"00Q{i:06}",
            "Name": f"Lead {i}",
            "Score": i * 3.5,
            "Active": (i % 2 == 0),
        }
        for i in range(100)
    ]
    path = str(tmp_path / "round_trip.xlsx")

    write_xlsx_streaming(rows, path, sheet_name="Leads")

    assert os.path.exists(path)
    read_back = _read_xlsx(path)
    assert len(read_back) == 100
    # Headers preserved in insertion order.
    assert list(read_back[0].keys()) == ["Id", "Name", "Score", "Active"]
    # Sample a few rows to spot-check round-trip fidelity.
    assert read_back[0]["Id"] == "00Q000000"
    assert read_back[7]["Name"] == "Lead 7"
    assert read_back[42]["Score"] == 42 * 3.5
    # Booleans round-trip as booleans (openpyxl preserves the type).
    assert read_back[10]["Active"] is True
    assert read_back[11]["Active"] is False


def test_write_xlsx_handles_extra_and_missing_keys(tmp_path: Path):
    """Rows with extra/missing keys against the first row's header set."""
    rows = [
        {"a": 1, "b": 2},
        {"a": 3, "c": 99},  # missing 'b', extra 'c'
        {"a": 5},  # missing 'b'
    ]
    path = str(tmp_path / "shape.xlsx")
    write_xlsx_streaming(rows, path)

    read_back = _read_xlsx(path)
    assert list(read_back[0].keys()) == ["a", "b"]
    assert read_back[1]["b"] in ("", None)  # missing → blank
    assert read_back[2]["b"] in ("", None)
    # Extra 'c' on row 1 must not have created a column.
    assert "c" not in read_back[1]


def test_write_xlsx_empty_list_writes_valid_file(tmp_path: Path):
    """Empty rows still writes a valid (empty) workbook. Direct call only —
    virtualize_result short-circuits before reaching here."""
    path = str(tmp_path / "empty.xlsx")
    write_xlsx_streaming([], path)
    assert os.path.exists(path)


# ──────────────────────────────────────────────────────────────────────────
# compute_summary_stats — per-column shape
# ──────────────────────────────────────────────────────────────────────────


def test_summary_stats_numeric_column():
    """Numeric column reports min, max, mean, count, dtype, null_count."""
    rows = [{"score": 10}, {"score": 20}, {"score": 30}, {"score": None}]
    stats = compute_summary_stats(rows)
    s = stats["score"]
    assert s["dtype"] == "int"
    assert s["min"] == 10
    assert s["max"] == 30
    assert s["mean"] == 20.0
    assert s["count"] == 3
    assert s["null_count"] == 1


def test_summary_stats_low_cardinality_string_column():
    """Low-cardinality string column reports top_values (sorted by count)."""
    rows = [
        {"stage": "Open"},
        {"stage": "Open"},
        {"stage": "Open"},
        {"stage": "Closed Won"},
        {"stage": "Closed Won"},
        {"stage": "Lost"},
    ]
    stats = compute_summary_stats(rows)
    s = stats["stage"]
    assert s["dtype"] == "str"
    assert "top_values" in s
    # Top entry is the most frequent.
    assert s["top_values"][0] == {"value": "Open", "count": 3}
    assert s["top_values"][1] == {"value": "Closed Won", "count": 2}
    assert s["top_values"][2] == {"value": "Lost", "count": 1}


def test_summary_stats_high_cardinality_string_column():
    """High-cardinality string column reports unique_count, not top_values."""
    rows = [{"email": f"user{i}@x.com"} for i in range(LOW_CARDINALITY_MAX + 5)]
    stats = compute_summary_stats(rows)
    s = stats["email"]
    assert s["dtype"] == "str"
    assert "top_values" not in s
    assert s["unique_count"] == LOW_CARDINALITY_MAX + 5
    assert s["null_count"] == 0


def test_summary_stats_bool_column():
    """Boolean column gets true_count + false_count, not top_values."""
    rows = [{"active": True}] * 7 + [{"active": False}] * 3
    stats = compute_summary_stats(rows)
    s = stats["active"]
    assert s["dtype"] == "bool"
    assert s["true_count"] == 7
    assert s["false_count"] == 3


def test_summary_stats_empty_input():
    """Empty rows returns an empty dict."""
    assert compute_summary_stats([]) == {}


# ──────────────────────────────────────────────────────────────────────────
# virtualize_result — public entrypoint
# ──────────────────────────────────────────────────────────────────────────


def test_virtualize_zero_rows_no_file_written(tmp_path: Path):
    """0-row input returns the full shape with file_path=None and no file."""
    result = virtualize_result([], tool_name="soqlQuery", output_dir=str(tmp_path))

    assert result["row_count"] == 0
    assert result["preview"] == []
    assert result["file_path"] is None
    assert result["summary_stats"] == {}
    assert result["schema"] == {}
    assert "next_steps" in result

    # No files written.
    assert list(tmp_path.iterdir()) == []


def test_virtualize_returns_preview_and_file_path(tmp_path: Path):
    """Above PREVIEW_ROW_COUNT input returns 10-row preview + writes file."""
    rows = [{"id": i, "stage": "Open" if i % 2 else "Closed"} for i in range(25)]
    result = virtualize_result(rows, tool_name="soqlQuery", output_dir=str(tmp_path))

    assert result["row_count"] == 25
    assert len(result["preview"]) == PREVIEW_ROW_COUNT
    assert result["preview"][0] == {"id": 0, "stage": "Closed"}
    assert result["file_path"] is not None
    assert os.path.exists(result["file_path"])
    # Schema flattens the per-column dtype.
    assert result["schema"] == {"id": "int", "stage": "str"}
    # Summary stats present for both columns.
    assert result["summary_stats"]["id"]["count"] == 25
    assert "top_values" in result["summary_stats"]["stage"]
    # File name includes tool name and is in the output_dir.
    assert "soqlQuery" in os.path.basename(result["file_path"])


def test_virtualize_filename_sanitizes_tool_name(tmp_path: Path):
    """Tool names with funky chars get sanitized to a safe filename component."""
    rows = [{"a": 1}]
    result = virtualize_result(
        rows, tool_name="weird/tool name!", output_dir=str(tmp_path)
    )
    # No slashes, spaces, or punctuation in the basename.
    base = os.path.basename(result["file_path"])
    assert "/" not in base
    assert " " not in base
    assert "!" not in base


# ──────────────────────────────────────────────────────────────────────────
# Performance — 10K rows in <5s
# ──────────────────────────────────────────────────────────────────────────


def test_two_virtualizations_in_same_second_produce_distinct_files(tmp_path: Path):
    """Two rapid virtualizations from the same tool must not collide.

    Codex review PR #99 (comment 3223912577): the old filename was
    ``{tool}_{whole_seconds}.xlsx``, so two large queries that landed in
    the same second overwrote each other and silently corrupted any
    attachment handles still pointing at the first dataset. Microseconds
    plus a 4-char random hex suffix make collisions astronomically unlikely.
    """
    rows_a = [{"id": i, "label": "A"} for i in range(3)]
    rows_b = [{"id": i, "label": "B"} for i in range(3)]

    result_a = virtualize_result(
        rows_a, tool_name="soqlQuery", output_dir=str(tmp_path)
    )
    result_b = virtualize_result(
        rows_b, tool_name="soqlQuery", output_dir=str(tmp_path)
    )

    # Different filenames.
    assert result_a["file_path"] != result_b["file_path"], (
        f"same-second virtualizations collided: {result_a['file_path']!r}"
    )
    # Both files exist on disk.
    assert os.path.exists(result_a["file_path"])
    assert os.path.exists(result_b["file_path"])
    # Each file has its own dataset — neither was overwritten.
    rows_a_read = _read_xlsx(result_a["file_path"])
    rows_b_read = _read_xlsx(result_b["file_path"])
    assert all(r["label"] == "A" for r in rows_a_read)
    assert all(r["label"] == "B" for r in rows_b_read)


def test_virtualize_10k_rows_under_5_seconds(tmp_path: Path):
    """10K-row pull must complete in <5s. Streaming write keeps memory bounded."""
    rows = [
        {
            "Id": f"00Q{i:08}",
            "Name": f"Lead {i}",
            "Stage": ("Open", "Closed Won", "Lost", "Qualified")[i % 4],
            "Score": i * 0.5,
        }
        for i in range(10_000)
    ]

    start = time.monotonic()
    result = virtualize_result(rows, tool_name="soqlQuery", output_dir=str(tmp_path))
    elapsed = time.monotonic() - start

    assert result["row_count"] == 10_000
    assert os.path.exists(result["file_path"])
    assert elapsed < 5.0, f"virtualize_result took {elapsed:.2f}s — should be <5s"


# ---------------------------------------------------------------------------
# Regression — nested SF relationship fields (Owner.Name, RecordType.Name)
# arrive as OrderedDicts. The summary stats layer must coerce them into
# hashable values, NOT raise ``unhashable type: 'collections.OrderedDict'``.
# Live repro 2026-05-14: session sesn_EXAMPLE, dump_sf_query
# failed 3+ times on Opportunity rows whose Owner / RecordType were nested.
# ---------------------------------------------------------------------------


def test_summary_stats_handles_ordered_dict_relationship_fields():
    """OrderedDict values from SF relationship traversal must not crash stats."""
    from collections import OrderedDict

    rows = [
        {
            "Id": "006abc",
            "Name": "Opp 1",
            "Owner": OrderedDict([("Name", "Alice"), ("Id", "005x")]),
            "RecordType": OrderedDict([("Name", "Direct")]),
        },
        {
            "Id": "006def",
            "Name": "Opp 2",
            "Owner": OrderedDict([("Name", "Alice"), ("Id", "005x")]),
            "RecordType": OrderedDict([("Name", "Channel")]),
        },
        {
            "Id": "006ghi",
            "Name": "Opp 3",
            "Owner": OrderedDict([("Name", "Bob"), ("Id", "005y")]),
            "RecordType": OrderedDict([("Name", "Direct")]),
        },
    ]
    # Before the fix this raised TypeError. After: nested dicts collapse into
    # JSON strings that the set / counter can hash.
    out = compute_summary_stats(rows)

    # Both nested columns made it through and got top_values.
    assert "Owner" in out
    assert "RecordType" in out
    # Two distinct Owners (Alice ×2, Bob ×1).
    owner_top = out["Owner"]["top_values"]
    assert len(owner_top) == 2
    # Alice's JSON-coerced value appears first (count 2).
    assert '"Alice"' in str(owner_top[0]["value"])
    assert owner_top[0]["count"] == 2
    # Two distinct RecordTypes (Direct ×2, Channel ×1).
    rt_top = out["RecordType"]["top_values"]
    assert len(rt_top) == 2


def test_top_values_handles_list_of_lists():
    """Lists are coerced too — fan-out queries can return tag arrays."""
    rows = [
        {"tags": ["a", "b"]},
        {"tags": ["a", "b"]},
        {"tags": ["a", "c"]},
    ]
    # Doesn't matter what stats says about the row (the dtype heuristic may
    # call this "str" or fall through); the test is that it doesn't raise.
    compute_summary_stats(rows)
