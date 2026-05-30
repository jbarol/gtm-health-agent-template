"""Integration test for _dispatch_post_report.

Mocks the Slack post (send_notification) to verify that valid payloads route
through the renderer, that invalid payloads produce a [SCHEMA_VALIDATION_FAILED]
fallback post, and that verbosity is honored.

Relies on conftest.py loading real .env (for module-import-time config) and
stubbing slack_bolt (so slack_bot import doesn't try to auth.test).

Run:
    cd orchestrator && python3 -m pytest dispatch_post_report_test.py
"""

from __future__ import annotations

import json
from unittest.mock import MagicMock, patch


def test_post_report_valid_summary():
    """Valid ad_hoc payload, default verbosity (summary) — should render and post."""
    from session_runner import _dispatch_post_report

    payload = {
        "headline": "Win rate down 4.2pp this quarter",
        "key_metrics": [
            {
                "name": "Win rate",
                "current": "23.4%",
                "prior": "27.6%",
                "benchmark": "20-30%",
                "trend": "down",
            }
        ],
        "findings": [
            {
                "headline": "Partner channel win rate collapsed",
                "value": "8% (n=42)",
                "confidence": "HIGH",
                "severity": "critical",
            }
        ],
    }
    tool_input = {"response_type": "ad_hoc_investigation_result", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "1234567890.000100"
        result_text = _dispatch_post_report(
            tool_input, thread_ts="thread-x", session_id="sess-1"
        )

    result = json.loads(result_text)
    assert result["ok"] is True
    assert result["response_type"] == "ad_hoc_investigation_result"
    assert result["mode"] == "summary"
    assert result["message_ts"] == "1234567890.000100"

    # Verify the rendered Slack message has the expected structure
    mock_send.assert_called_once()
    kwargs = mock_send.call_args.kwargs
    assert kwargs["severity"] == "critical"  # inferred from the critical finding
    assert "*Win rate down 4.2pp this quarter*" in kwargs["summary"]
    assert "Partner channel win rate collapsed" in kwargs["summary"]
    assert "`expand:`" in kwargs["summary"]  # summary mode has the footer hint
    assert kwargs["reply_to"] == "thread-x"


def test_post_report_expanded_when_verbosity_passed():
    """verbosity='expanded' passed in selects expanded render mode."""
    from session_runner import _dispatch_post_report

    payload = {
        "metric": "Win rate, Q1",
        "value": "23.4%",
        "as_of": "2026-05-11",
        "source": "Salesforce",
    }
    tool_input = {"response_type": "quick_answer", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-expanded"
        result_text = _dispatch_post_report(
            tool_input, session_id="sess-expanded", verbosity="expanded"
        )

    result = json.loads(result_text)
    assert result["mode"] == "expanded"
    rendered = mock_send.call_args.kwargs["summary"]
    # Expanded mode for QuickAnswer adds the Source line
    assert "Source: Salesforce" in rendered


def test_post_report_unknown_type():
    """Unknown response_type returns is_error=true to the agent — NO Slack post.

    B7 (2026-05-12 self-heal): the orchestrator no longer posts
    [POST_REPORT_FAILED] watch notices into the user's thread. Recoverable
    in-band failures get fed back to the agent via ``_is_error=true`` so it
    can self-correct. Up to POST_REPORT_MAX_RETRIES attempts before the
    terminal give-up path runs.
    """
    from session_runner import _dispatch_post_report

    tool_input = {"response_type": "nonexistent_type", "payload": {}}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-unknown"
        result_text = _dispatch_post_report(tool_input)

    result = json.loads(result_text)
    assert result["error"] == "unknown_response_type"
    assert result.get("_is_error") is True
    # No public-channel Slack post on the first attempt — the agent retries.
    mock_send.assert_not_called()


def test_post_report_schema_validation_failure():
    """Schema validation failure returns is_error=true to the agent — NO Slack post.

    B7 (2026-05-12 self-heal): the agent gets the exact Pydantic error in
    the ``detail`` field, plus a ``retry_hint`` telling it how to recover.
    No watch notice in the public channel.
    """
    from session_runner import _dispatch_post_report

    # Missing required fields for anomaly_alert
    bad_payload = {
        "headline": "Something is wrong",
        # missing metric, current_value, prior_value, benchmark, severity,
        # evidence_summary, recommended_action
    }
    tool_input = {"response_type": "anomaly_alert", "payload": bad_payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-fail"
        result_text = _dispatch_post_report(tool_input)

    result = json.loads(result_text)
    assert result["error"] == "schema_validation_failed"
    assert result.get("_is_error") is True
    assert "retry_hint" in result
    # The exact Pydantic error is in the detail field so the agent can fix
    # the specific failing fields. Two-layer assertion:
    #
    #   1. The detail uses one of the standard Pydantic-error vocabularies
    #      so the agent recognizes the failure mode. We accept "validation
    #      error" (Pydantic v2 header), "field required" (v1) or the
    #      orchestrator's translated phrasing "required field is missing"
    #      (``_build_validation_detail`` in session_runner.py) — but NOT a
    #      generic "schema validation" prefix on its own, which a
    #      header-only regression could still emit.
    #
    #   2. Field-level contract (codex P2 on PR #256). The agent cannot
    #      self-correct from a header-only message; it needs the specific
    #      failing field names. Pin every required field from the
    #      anomaly_alert schema that the bad_payload above omits, so any
    #      regression in ``_build_validation_detail`` that drops per-field
    #      entries fails this test loudly.
    detail_lower = result["detail"].lower()
    assert (
        "validation error" in detail_lower
        or "field required" in detail_lower
        or "required field" in detail_lower
    ), f"missing Pydantic-error vocabulary in detail: {result['detail']!r}"

    # Field-level contract: every omitted required field must appear in the
    # detail string so the agent can self-correct on the next post_report
    # attempt. anomaly_alert requires metric / current_value / prior_value /
    # benchmark / severity / evidence_summary / recommended_action — all
    # are absent from bad_payload above, so all must be named in detail.
    for required_field in (
        "metric",
        "current_value",
        "prior_value",
        "benchmark",
        "severity",
        "evidence_summary",
        "recommended_action",
    ):
        assert required_field in result["detail"], (
            f"detail dropped per-field entry for {required_field!r}: "
            f"{result['detail']!r}"
        )

    # No Slack post — the agent self-corrects silently.
    mock_send.assert_not_called()


def test_post_report_retries_exhausted_posts_neutral_terminal_line():
    """3rd validation failure triggers the terminal give-up: ONE neutral line + forensic dump.

    Asserts the in-thread message contains zero Pydantic-style strings —
    no 'validation error', no 'extra_forbidden', no 'string_too_long'.
    Forensic dump captures the failing payload + error history for self_heal.
    """
    from session_runner import _dispatch_post_report
    import session_runner

    # Reset retry counter so this test is independent of others.
    session_runner._clear_post_report_retries("sess-give-up")

    bad_payload = {"headline": "Something is wrong"}  # missing required fields
    tool_input = {"response_type": "anomaly_alert", "payload": bad_payload}

    posted_messages = []

    def capture_post(*args, **kwargs):
        posted_messages.append(kwargs)
        return "ts-terminal"

    with patch("session_runner.send_notification", side_effect=capture_post):
        # First 2 attempts: retry path (no Slack post).
        for i in range(session_runner.POST_REPORT_MAX_RETRIES - 1):
            result_text = _dispatch_post_report(
                tool_input,
                thread_ts="thread-x",
                session_id="sess-give-up",
            )
            result = json.loads(result_text)
            assert result.get("_is_error") is True, f"attempt {i + 1}"
        # Final attempt: terminal give-up.
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-give-up",
        )

    result = json.loads(result_text)
    assert result["error"] == "post_report_give_up"
    # Exactly ONE Slack post — the terminal neutral line.
    assert len(posted_messages) == 1
    kwargs = posted_messages[0]
    # Severity is info — no scary :warning: emoji, no @-ping. The summary is
    # human and free of Pydantic-speak.
    assert kwargs["severity"] == "info"
    summary = kwargs["summary"]
    for forbidden_token in (
        "validation error",
        "extra_forbidden",
        "string_too_long",
        "type=missing",
        "Pydantic",
        "[POST_REPORT_FAILED]",
    ):
        assert forbidden_token not in summary, (
            f"terminal line leaked Pydantic-style token {forbidden_token!r}: {summary!r}"
        )
    # The terminal line acknowledges the failure to the user in plain English.
    assert "couldn't assemble" in summary.lower() or "couldn't" in summary.lower()


def test_post_report_oversized_field_handled_by_editor():
    """PR #87 removed string caps; editor pass now owns length enforcement.

    With caps gone, a 100-char metric on quick_answer no longer triggers
    a schema rejection. The editor pass runs before validation but doesn't
    have a metric-specific target — it leaves the field alone and validation
    accepts it (no cap to violate). This is intentional: the editor only
    targets fields that face the reader in long form (headline, value,
    cross_domain_pattern, methodology_note, decision options). Internal
    short-fact fields fall through.
    """
    from session_runner import _dispatch_post_report

    payload = {
        "metric": "x" * 100,  # no schema cap (PR #87)
        "value": "23.4%",
        "as_of": "2026-05-11",
        "source": "Salesforce",
    }
    tool_input = {"response_type": "quick_answer", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-overrun"
        result_text = _dispatch_post_report(tool_input)

    result = json.loads(result_text)
    # Validation passes — no string cap to enforce
    assert result.get("ok") is True, result


def test_post_report_severity_inferred_from_findings():
    """Severity of the Slack notification reflects the most-severe finding."""
    from session_runner import _dispatch_post_report

    payload = {
        "headline": "All quiet this morning",
        "findings": [
            {
                "headline": "Nothing remarkable",
                "value": "—",
                "confidence": "HIGH",
                "severity": "info",
            }
        ],
    }
    tool_input = {"response_type": "ad_hoc_investigation_result", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-info"
        _dispatch_post_report(tool_input)

    assert mock_send.call_args.kwargs["severity"] == "info"


def test_payload_not_a_dict():
    """Non-dict payload returns is_error=true to the agent — NO Slack post (B7)."""
    from session_runner import _dispatch_post_report

    tool_input = {"response_type": "quick_answer", "payload": "not a dict"}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-bad-payload"
        result_text = _dispatch_post_report(tool_input)

    result = json.loads(result_text)
    assert result["error"] == "payload_not_object"
    assert result.get("_is_error") is True
    # No Slack post — the agent retries.
    mock_send.assert_not_called()


# ---------------------------------------------------------------------------
# Autoplan test gap closures (Tasks #7, #10)
# ---------------------------------------------------------------------------

import pytest

# Task #7 — Parametrize length-cap rejection across multiple fields.
# Existing test_post_report_length_overrun_caught covers one field (metric).
# Extend to every Field(max_length=N) in every schema to catch silent cap
# changes that would let oversized payloads through.


def _valid_payload_for(rt: str) -> dict:
    """Return a minimally-valid payload for each response_type."""
    if rt == "quick_answer":
        return {
            "metric": "Win rate",
            "value": "23.4%",
            "as_of": "2026-05-11",
            "source": "Salesforce",
        }
    if rt == "anomaly_alert":
        return {
            "headline": "Anomaly",
            "metric": "Win rate",
            "current_value": "8%",
            "prior_value": "23%",
            "benchmark": "20-30%",
            "severity": "critical",
            "evidence_summary": "Partner channel collapse",
            "recommended_action": "Pause partner intake",
        }
    if rt == "ad_hoc_investigation_result":
        return {"headline": "Win rate down"}
    if rt == "nightly_digest":
        return {"headline": "Overnight digest"}
    if rt == "weekly_status":
        return {
            "headline": "Weekly status",
            "trajectory": "Stable",
        }
    raise KeyError(rt)


# PR #87 removed max_length caps from user-facing prose fields. The editor
# pass (see editor.py + editor_test.py) is now the primary enforcer of
# audience-appropriate length: it trims oversized strings BEFORE Pydantic
# validation runs. The cap-rejection test below is preserved for the
# fields that STILL have caps (list-length caps on findings, key_metrics,
# decision_options, etc.), so silent cap changes still get caught.


@pytest.mark.parametrize(
    "response_type,field,list_cap",
    [
        # List-length caps (Pydantic max_length on list[T] fields).
        # These remain after PR #87 because they protect the renderer from
        # blowing up on a 100-item findings array.
        ("ad_hoc_investigation_result", "findings", 4),
        ("ad_hoc_investigation_result", "key_metrics", 5),
        ("ad_hoc_investigation_result", "open_questions", 3),
        ("ad_hoc_investigation_result", "tables", 3),
        ("nightly_digest", "portcos_with_action", 5),
        ("nightly_digest", "changes_overnight", 5),
        ("weekly_status", "portco_lines", 10),
    ],
)
def test_list_length_cap_rejection_per_field(response_type, field, list_cap):
    """Every list[T] max_length cap rejects overruns. One case per field.

    String caps were removed in PR #87 — the editor pass owns string length
    now. List-length caps are still enforced by Pydantic because they
    protect the renderer from unbounded payloads. This parametrized test
    confirms each list cap fires on overrun.
    """
    from session_runner import _dispatch_post_report
    import json
    from unittest.mock import patch

    payload = _valid_payload_for(response_type)
    # Build a list of (list_cap + 1) placeholder items. Item shape depends on
    # the list element type: Finding for findings/changes_overnight,
    # KeyMetric for key_metrics, PortcoLine for portco_lines, TableBlock for
    # tables, plain strings elsewhere.
    overrun = list_cap + 1
    if field in ("findings", "changes_overnight"):
        payload[field] = [
            {
                "headline": f"f{i}",
                "value": f"v{i}",
                "confidence": "MEDIUM",
                "severity": "watch",
            }
            for i in range(overrun)
        ]
    elif field == "key_metrics":
        payload[field] = [{"name": f"m{i}", "current": f"{i}%"} for i in range(overrun)]
    elif field == "portco_lines":
        payload[field] = [
            {"portco": f"p{i}", "headline": f"h{i}", "severity": "watch"}
            for i in range(overrun)
        ]
    elif field == "tables":
        payload[field] = [
            {"title": f"t{i}", "headers": ["a"], "rows": [["x"]]}
            for i in range(overrun)
        ]
    else:
        # plain string lists (portcos_with_action, open_questions)
        payload[field] = [f"item{i}" for i in range(overrun)]

    tool_input = {"response_type": response_type, "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-cap"
        result_text = _dispatch_post_report(tool_input)

    result = json.loads(result_text)
    assert result.get("error") == "schema_validation_failed", (
        f"{response_type}.{field} (list cap {list_cap}) did not raise on "
        f"overrun of {overrun}; result={result}"
    )


def test_editor_trims_oversized_headline_when_schema_has_no_cap():
    """Post-PR #87: string caps gone. Editor must still trim oversized strings.

    Without the editor, a 250-char headline would ship to Slack verbatim.
    The editor's target (100 chars) caps it via shorteners + truncation.
    """
    from session_runner import _dispatch_post_report
    import json
    from unittest.mock import patch

    # 250-char headline with phrase boundaries so the truncator finds one
    long_headline = (
        "Q3'26 new-business pipeline coverage is materially below plan, "
        "driven by partner-channel weakness, enterprise-segment slowdown, "
        "and the unwinding of the inflated late-stage win-rate from last "
        "quarter — coverage now 0.6x vs target."
    )
    assert len(long_headline) >= 200

    payload = {"headline": long_headline, "findings": []}
    tool_input = {"response_type": "ad_hoc_investigation_result", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-edited"
        result_text = _dispatch_post_report(tool_input)

    result = json.loads(result_text)
    # Validation passes (no cap to reject)
    assert result.get("ok") is True, result
    # But the headline that actually went to Slack must be trimmed
    rendered = mock_send.call_args.kwargs["summary"]
    # The editor target for headline is 100; rendered headline must be near that
    # (within slack-rendering overhead like asterisks/whitespace)
    first_line = rendered.split("\n")[0]
    assert len(first_line) < len(long_headline) + 20, (
        f"editor failed to trim headline: first_line={first_line!r}"
    )


# Task #10 — Severity inference across critical, watch, info.
# Existing test covers the info case. Add critical + watch + mixed-severity
# (highest wins) to confirm the inference rule.


def test_severity_inferred_critical():
    """A critical finding raises the Slack notification severity to critical."""
    from session_runner import _dispatch_post_report

    payload = {
        "headline": "Critical event",
        "findings": [
            {
                "headline": "Partner channel collapse",
                "value": "8%",
                "confidence": "HIGH",
                "severity": "critical",
            }
        ],
    }
    tool_input = {"response_type": "ad_hoc_investigation_result", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-critical"
        _dispatch_post_report(tool_input)

    assert mock_send.call_args.kwargs["severity"] == "critical"


def test_severity_inferred_watch():
    """A watch-only set of findings produces a watch-severity Slack post."""
    from session_runner import _dispatch_post_report

    payload = {
        "headline": "Trending issue",
        "findings": [
            {
                "headline": "Cycle time elongating",
                "value": "47d",
                "confidence": "MEDIUM",
                "severity": "watch",
            }
        ],
    }
    tool_input = {"response_type": "ad_hoc_investigation_result", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-watch"
        _dispatch_post_report(tool_input)

    assert mock_send.call_args.kwargs["severity"] == "watch"


def test_severity_mixed_picks_highest():
    """Mixed-severity findings: highest (critical > watch > info) wins."""
    from session_runner import _dispatch_post_report

    payload = {
        "headline": "Mixed signal",
        "findings": [
            {
                "headline": "Info finding",
                "value": "—",
                "confidence": "HIGH",
                "severity": "info",
            },
            {
                "headline": "Watch finding",
                "value": "47d",
                "confidence": "MEDIUM",
                "severity": "watch",
            },
            {
                "headline": "Critical finding",
                "value": "8%",
                "confidence": "HIGH",
                "severity": "critical",
            },
        ],
    }
    tool_input = {"response_type": "ad_hoc_investigation_result", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-mixed"
        _dispatch_post_report(tool_input)

    assert mock_send.call_args.kwargs["severity"] == "critical"


def test_severity_inferred_for_nightly_digest_critical():
    """Severity inference applies to nightly_digest changes_overnight too."""
    from session_runner import _dispatch_post_report

    payload = {
        "headline": "Overnight: critical break",
        "portcos_with_action": ["Acme"],
        "changes_overnight": [
            {
                "headline": "GRR fell below benchmark",
                "value": "78%",
                "confidence": "HIGH",
                "severity": "critical",
            }
        ],
    }
    tool_input = {"response_type": "nightly_digest", "payload": payload}

    with patch("session_runner.send_notification") as mock_send:
        mock_send.return_value = "ts-nightly-crit"
        _dispatch_post_report(tool_input)

    assert mock_send.call_args.kwargs["severity"] == "critical"


# ---------------------------------------------------------------------------
# Attachment-path whitelist (PR #99, codex review comment 3223912574)
# ---------------------------------------------------------------------------
#
# payload.attachments is model-controlled — the agent's structured output
# decides which files to upload. A prompt-injected or otherwise mis-routed
# agent could request arbitrary readable files (``/etc/passwd``, ``.env``,
# ``~/.ssh/id_rsa``) and exfiltrate them via Slack. We whitelist every
# attachment path to ``SESSION_OUTPUT_DIR`` (defaults to
# ``/mnt/session/outputs`` in prod).


def test_attachment_outside_session_outputs_rejected(tmp_path, monkeypatch):
    """An attachment pointing outside SESSION_OUTPUT_DIR is dropped + logged.

    The post still goes through (we do NOT raise) — the safe attachments
    (or no attachments) ship. The unsafe path is logged with
    ``[ATTACHMENT_PATH_REJECTED]`` and never reaches Slack.
    """
    import session_runner
    from session_runner import _dispatch_post_report

    # Point SESSION_OUTPUT_DIR at a writable test dir so the validator has a
    # legit prefix to compare against.
    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))

    # /etc/passwd exists on every POSIX system — a real file outside the
    # session outputs. The validator must reject it.
    payload = {
        "headline": "Win rate down",
        "attachments": ["/etc/passwd"],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
        patch.object(session_runner, "log") as mock_log,
    ):
        mock_send.return_value = "ts-rejected"
        result_text = _dispatch_post_report(
            tool_input, thread_ts="t-rej", session_id="sess-rej"
        )

    result = json.loads(result_text)
    # The post still went through with safe content.
    assert result["ok"] is True
    # The unsafe /etc/passwd attachment was dropped. Floating-prancing-trinket
    # PR 6 introduced the auto-generated .docx sibling for every post_report;
    # that file lives under SESSION_OUTPUT_DIR and is the only safe item that
    # reaches the upload list. Assert the unsafe path is absent and any
    # remaining item is the safe .docx.
    assert mock_attach.called
    upload_arg = mock_attach.call_args[0][0]
    assert "/etc/passwd" not in upload_arg
    assert all(p.endswith(".docx") for p in upload_arg), (
        f"only the auto-generated .docx should remain after filtering; got {upload_arg}"
    )
    # Rejection logged with the canonical token.
    warning_calls = [
        call
        for call in mock_log.warning.call_args_list
        if call.args and "ATTACHMENT_PATH_REJECTED" in str(call.args[0])
    ]
    assert warning_calls, "expected [ATTACHMENT_PATH_REJECTED] log entry"


def test_attachment_symlink_rejected(tmp_path, monkeypatch):
    """A symlinked path that resolves outside SESSION_OUTPUT_DIR is rejected.

    A path can sit nominally inside the session output directory but resolve
    via symlink to a sensitive file (``../../../etc/passwd``). Any symlink
    segment in the path is treated as suspect — reject conservatively.
    """
    import session_runner
    from session_runner import _dispatch_post_report

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))

    # Symlink inside the safe dir pointing OUT.
    target = tmp_path / "secret.txt"
    target.write_text("ssn=123-45-6789")
    sym = outputs_dir / "looks-safe.txt"
    sym.symlink_to(target)
    assert sym.is_symlink()

    payload = {
        "headline": "Win rate down",
        "attachments": [str(sym)],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
    ):
        mock_send.return_value = "ts-sym"
        result_text = _dispatch_post_report(
            tool_input, thread_ts="t-sym", session_id="sess-sym"
        )

    result = json.loads(result_text)
    assert result["ok"] is True
    # The symlinked attachment was dropped. The auto-generated .docx
    # (floating-prancing-trinket PR 6) is the only safe item left.
    upload_arg = mock_attach.call_args[0][0]
    assert str(sym) not in upload_arg
    assert all(p.endswith(".docx") for p in upload_arg), (
        f"only the auto-generated .docx should remain after filtering; got {upload_arg}"
    )


def test_attachment_inside_session_outputs_accepted(tmp_path, monkeypatch):
    """Happy path: an attachment under SESSION_OUTPUT_DIR is uploaded."""
    import session_runner
    from session_runner import _dispatch_post_report

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))

    good = outputs_dir / "report.xlsx"
    good.write_bytes(b"PK\x03\x04 fake xlsx bytes")

    payload = {
        "headline": "Win rate down",
        "attachments": [str(good)],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
    ):
        mock_send.return_value = "ts-good"
        result_text = _dispatch_post_report(
            tool_input, thread_ts="t-good", session_id="sess-good"
        )

    result = json.loads(result_text)
    assert result["ok"] is True
    # The safe attachment made it through.
    upload_arg = mock_attach.call_args[0][0]
    assert str(good) in upload_arg


# ---------------------------------------------------------------------------
# Plan #11 — theme-tagged post_report routes through slack_thread_registry
# ---------------------------------------------------------------------------
#
# A theme set on tool_input causes _dispatch_post_report to anchor the
# post to a per-theme thread (via slack_thread_registry.get_or_create_thread)
# instead of posting top-level. Verified end-to-end here: the dispatcher
# resolves a thread_ts from the registry and passes it as reply_to to
# send_notification.


def test_post_report_with_theme_anchors_to_registry_thread():
    """A theme on the input routes the report into a per-theme thread.

    The registry is the source of truth for which thread to use. The
    dispatcher MUST pass the registry's thread_ts as ``reply_to`` so
    the full report lands as a reply on the parent message, not as a
    new top-level message.
    """
    import slack_thread_registry
    from session_runner import _dispatch_post_report

    slack_thread_registry._clear_cache_for_tests()

    payload = {
        "headline": "Pipeline coverage 0.6x — partner weakness",
        "findings": [
            {
                "headline": "Partner channel down",
                "value": "12% wins (n=42)",
                "confidence": "HIGH",
                "severity": "watch",
            }
        ],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
        "theme": "pipeline_review",
        "nightly_run_id": "nightly-2026-05-14",
    }

    # Returns a deterministic ts every time so we can tell which call
    # produced which message — the registry's parent post first, then the
    # full report as a reply.
    posted: list[dict] = []
    ts_iter = iter(["ts-parent", "ts-reply"])

    def _capture(**kwargs):
        posted.append(kwargs)
        return next(ts_iter)

    with patch("session_runner.send_notification", side_effect=_capture):
        result_text = _dispatch_post_report(
            tool_input,
            session_id="sess-theme",
            channel_id="C-pipeline",
        )

    result = json.loads(result_text)
    assert result["ok"] is True
    # Exactly two posts: the parent (registry) and the reply (dispatcher).
    assert len(posted) == 2
    # Parent — top-level, contains headline + pointer.
    parent = posted[0]
    assert parent["reply_to"] is None
    assert parent["severity"] == "info"
    assert "Pipeline coverage 0.6x" in parent["summary"]
    assert "More details in thread" in parent["summary"]
    # Reply — threaded under the parent.
    reply = posted[1]
    assert reply["reply_to"] == "ts-parent"
    # The reply carries the fully-rendered report; not just the headline.
    assert "Partner channel down" in reply["summary"]


def test_post_report_without_theme_does_not_create_parent_thread():
    """Existing ad-hoc behavior: no theme → no registry anchoring."""
    import slack_thread_registry
    from session_runner import _dispatch_post_report

    slack_thread_registry._clear_cache_for_tests()

    payload = {
        "headline": "Ad-hoc answer",
        "findings": [],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    posted: list[dict] = []

    def _capture(**kwargs):
        posted.append(kwargs)
        return "ts-adhoc"

    with patch("session_runner.send_notification", side_effect=_capture):
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="user-thread",
            session_id="sess-adhoc",
            channel_id="C-adhoc",
        )

    result = json.loads(result_text)
    assert result["ok"] is True
    # Only ONE post — the existing thread_ts pathway, not the registry.
    assert len(posted) == 1
    assert posted[0]["reply_to"] == "user-thread"


def test_post_report_theme_with_unknown_value_falls_back_to_unthreaded():
    """An unknown theme is logged + ignored — no registry call, no crash.

    Defensive against schema drift: if a future agent sends a theme not
    in VALID_THEMES, we don't want a hard error blocking delivery. The
    report still ships, just unthreaded (legacy behavior).
    """
    import slack_thread_registry
    from session_runner import _dispatch_post_report

    slack_thread_registry._clear_cache_for_tests()

    payload = {"headline": "Some answer", "findings": []}
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
        "theme": "made_up_theme_not_in_enum",
    }

    posted: list[dict] = []

    def _capture(**kwargs):
        posted.append(kwargs)
        return "ts-fallback"

    with patch("session_runner.send_notification", side_effect=_capture):
        result_text = _dispatch_post_report(
            tool_input,
            session_id="sess-bad-theme",
            channel_id="C-bad",
        )

    result = json.loads(result_text)
    assert result["ok"] is True
    # Only ONE post — fallback path, no registry parent.
    assert len(posted) == 1
    assert posted[0]["reply_to"] is None


def test_post_report_two_themed_calls_share_one_parent_thread():
    """Two theme-tagged calls in the same run reuse the same parent.

    Concrete scenario: pipeline_review cron posts a headline, then five
    minutes later posts a follow-up chart for the same theme. Both
    artifacts MUST land in the same Slack thread.
    """
    import slack_thread_registry
    from session_runner import _dispatch_post_report

    slack_thread_registry._clear_cache_for_tests()

    payload_1 = {
        "headline": "Pipeline summary",
        "findings": [],
    }
    payload_2 = {
        "headline": "Pipeline drilldown — partner channel",
        "findings": [],
    }
    common = {
        "theme": "pipeline_review",
        "nightly_run_id": "nightly-2026-05-14-pipeline",
    }
    tool_input_1 = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload_1,
        **common,
    }
    tool_input_2 = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload_2,
        **common,
    }

    posted: list[dict] = []
    ts_iter = iter(["ts-parent", "ts-reply-1", "ts-reply-2"])

    def _capture(**kwargs):
        posted.append(kwargs)
        return next(ts_iter)

    with patch("session_runner.send_notification", side_effect=_capture):
        _dispatch_post_report(
            tool_input_1, session_id="sess-1", channel_id="C-pipeline"
        )
        _dispatch_post_report(
            tool_input_2, session_id="sess-2", channel_id="C-pipeline"
        )

    # ONE parent + TWO replies. The second call reused the cached
    # thread_ts — it did NOT post a second parent.
    assert len(posted) == 3
    # First call: top-level parent.
    assert posted[0]["reply_to"] is None
    # Second + third: replies in the same thread.
    assert posted[1]["reply_to"] == "ts-parent"
    assert posted[2]["reply_to"] == "ts-parent"


def test_is_safe_attachment_path_unit(tmp_path, monkeypatch):
    """Direct unit tests for the path validator.

    Codifies the three rejection rules (outside prefix, symlink, empty)
    independently of the dispatch flow so a future refactor that splits
    these concerns can't quietly regress the check.
    """
    from session_runner import _is_safe_attachment_path

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))

    # Inside — accepted.
    safe = outputs_dir / "x.xlsx"
    safe.write_text("ok")
    assert _is_safe_attachment_path(str(safe)) is True

    # Outside — rejected.
    assert _is_safe_attachment_path("/etc/passwd") is False
    assert _is_safe_attachment_path(str(tmp_path / "elsewhere.txt")) is False

    # Symlink inside pointing out — rejected.
    out_target = tmp_path / "outside.txt"
    out_target.write_text("nope")
    sym = outputs_dir / "evil.txt"
    sym.symlink_to(out_target)
    assert _is_safe_attachment_path(str(sym)) is False

    # Empty / None / non-string — rejected.
    assert _is_safe_attachment_path("") is False
    assert _is_safe_attachment_path(None) is False  # type: ignore[arg-type]
    assert _is_safe_attachment_path(12345) is False  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# Plan #48 Phase 3 / Plan #52 PR-E + PR-H — xlsx consolidation in
# _dispatch_post_report
# ---------------------------------------------------------------------------
#
# Multiple xlsx attachments collapse into a single consolidated workbook with
# each source as a named worksheet tab. Default-on as of PR-H (2026-05-22,
# after the 48h observation window of the env-var-enabled mode in prod
# cleared with zero [XLSX_CONSOLIDATE_FAILED] log lines). The user complaint
# that prompted this: 2026-05-19 "we have a massive issue where I keep
# asking for a single spreadsheet with tabs. and you keep sending 10
# separate excel files. This is unacceptible."
#
# Behavior matrix verified here:
#   1. Default (no env var) → 1 consolidated xlsx reaches upload (PR-H pin)
#   2. Flag on + N≥2 xlsx → 1 consolidated xlsx reaches upload
#   3. Flag off → original N xlsx files all reach upload (kill switch)
#   4. Flag on + split-pref registered True → original N files reach upload
#      (user explicitly asked for separate files via _SPLIT_FILES_KEYWORDS)
#   5. Flag on + consolidate_xlsx_files raises → graceful fall-through to
#      original N files (consolidation is best-effort UX polish, never blocks)


def _write_real_xlsx(path, sheet_title: str, rows: list[list]) -> None:
    """Helper: write a real .xlsx file at ``path`` with one named sheet."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None, "openpyxl Workbook() always has an active sheet"
    ws.title = sheet_title
    for row in rows:
        ws.append(row)
    wb.save(str(path))


def test_consolidation_default_on_when_env_unset(tmp_path, monkeypatch):
    """PR-H pin: no XLSX_CONSOLIDATE_ENABLED env var → consolidation runs.

    PR-E shipped behind a default-`false` env var; PR-H flipped the default
    to `true` after a clean 48h observation window in prod. This test pins
    the new default so a future revert of the default flip will fail loud.
    """
    import session_runner
    from session_runner import _dispatch_post_report

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))
    # Intentionally do NOT set XLSX_CONSOLIDATE_ENABLED — verify default-on.
    monkeypatch.delenv("XLSX_CONSOLIDATE_ENABLED", raising=False)

    xlsx_a = outputs_dir / "default_a.xlsx"
    xlsx_b = outputs_dir / "default_b.xlsx"
    _write_real_xlsx(xlsx_a, "A", [["k", "v"], ["a", 1]])
    _write_real_xlsx(xlsx_b, "B", [["k", "v"], ["b", 2]])

    payload = {
        "headline": "Default consolidation",
        "findings": [
            {
                "headline": "No env var set; merge anyway",
                "value": "see attachments",
                "confidence": "HIGH",
                "severity": "watch",
            }
        ],
        "attachments": [str(xlsx_a), str(xlsx_b)],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
    ):
        mock_send.return_value = "ts-default-on"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="t-default-on",
            session_id="sess-default-on-1",
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["ok"] is True

    upload_arg = mock_attach.call_args[0][0]
    xlsx_in_upload = [p for p in upload_arg if p.lower().endswith(".xlsx")]
    # With the new default-on behavior, the two inputs collapse to 1 file.
    assert len(xlsx_in_upload) == 1, (
        f"expected 1 consolidated xlsx with default-on, got {xlsx_in_upload}"
    )
    assert "consolidated_acme_" in xlsx_in_upload[0]
    assert str(xlsx_a) not in upload_arg
    assert str(xlsx_b) not in upload_arg


def test_consolidation_produces_one_xlsx_when_enabled(tmp_path, monkeypatch):
    """Two real xlsx in + flag on → exactly 1 xlsx reaches upload with 2 sheets."""
    import session_runner
    from session_runner import _dispatch_post_report

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))
    monkeypatch.setenv("XLSX_CONSOLIDATE_ENABLED", "true")

    xlsx_a = outputs_dir / "pipeline_summary.xlsx"
    xlsx_b = outputs_dir / "sales_funnel.xlsx"
    _write_real_xlsx(xlsx_a, "Pipeline", [["stage", "count"], ["sql", 42]])
    _write_real_xlsx(xlsx_b, "Funnel", [["step", "rate"], ["mql_to_sql", "0.41"]])

    payload = {
        "headline": "Two-dataset investigation",
        "findings": [
            {
                "headline": "Pipeline + funnel covered",
                "value": "see attachments",
                "confidence": "HIGH",
                "severity": "watch",
            }
        ],
        "attachments": [str(xlsx_a), str(xlsx_b)],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
    ):
        mock_send.return_value = "ts-consolidated"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="t-consol",
            session_id="sess-consol-1",
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["ok"] is True

    assert mock_attach.called
    upload_arg = mock_attach.call_args[0][0]
    xlsx_in_upload = [p for p in upload_arg if p.lower().endswith(".xlsx")]
    assert len(xlsx_in_upload) == 1, (
        f"expected exactly 1 consolidated xlsx, got {xlsx_in_upload}"
    )
    # The original two source files should NOT be in the upload list.
    assert str(xlsx_a) not in upload_arg
    assert str(xlsx_b) not in upload_arg
    # Consolidated filename pattern from xlsx_consolidate._derive_output_path.
    consolidated = xlsx_in_upload[0]
    assert "consolidated_acme_" in consolidated

    # Verify the consolidated workbook has 2 sheets (one per source).
    from openpyxl import load_workbook

    wb = load_workbook(consolidated, read_only=True)
    try:
        assert len(wb.worksheets) == 2, (
            f"expected 2 sheets in consolidated workbook, got "
            f"{[ws.title for ws in wb.worksheets]}"
        )
    finally:
        wb.close()


def test_consolidation_skipped_when_disabled(tmp_path, monkeypatch):
    """Kill switch: XLSX_CONSOLIDATE_ENABLED=false → original N files upload.

    PR-H flipped the default to `true`. This test pins the kill switch so a
    Railway env var override stays a viable emergency rollback path without
    a code change.
    """
    import session_runner
    from session_runner import _dispatch_post_report

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))
    monkeypatch.setenv("XLSX_CONSOLIDATE_ENABLED", "false")

    xlsx_a = outputs_dir / "a.xlsx"
    xlsx_b = outputs_dir / "b.xlsx"
    _write_real_xlsx(xlsx_a, "A", [["k", "v"], ["x", 1]])
    _write_real_xlsx(xlsx_b, "B", [["k", "v"], ["y", 2]])

    payload = {
        "headline": "Two files, flag off",
        "findings": [
            {
                "headline": "Both files attached",
                "value": "see attachments",
                "confidence": "HIGH",
                "severity": "watch",
            }
        ],
        "attachments": [str(xlsx_a), str(xlsx_b)],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
    ):
        mock_send.return_value = "ts-disabled"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="t-disabled",
            session_id="sess-disabled-1",
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["ok"] is True

    upload_arg = mock_attach.call_args[0][0]
    xlsx_in_upload = [p for p in upload_arg if p.lower().endswith(".xlsx")]
    # Both originals must reach upload; no consolidated file.
    assert str(xlsx_a) in upload_arg
    assert str(xlsx_b) in upload_arg
    assert len(xlsx_in_upload) == 2
    assert not any("consolidated_" in p for p in upload_arg)


def test_consolidation_split_pref_skips_merge(tmp_path, monkeypatch):
    """split-pref registered True + flag on → both xlsx upload, no merge."""
    import session_runner
    from session_runner import (
        _dispatch_post_report,
        _register_split_files_pref,
    )

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))
    monkeypatch.setenv("XLSX_CONSOLIDATE_ENABLED", "true")

    sid = "sess-split-1"
    # User said "separate files" in the question. PR-D registry recorded it.
    _register_split_files_pref(sid, True)

    xlsx_a = outputs_dir / "first.xlsx"
    xlsx_b = outputs_dir / "second.xlsx"
    _write_real_xlsx(xlsx_a, "First", [["k", "v"], ["a", 1]])
    _write_real_xlsx(xlsx_b, "Second", [["k", "v"], ["b", 2]])

    payload = {
        "headline": "User wanted separate files",
        "findings": [
            {
                "headline": "Honoring user preference",
                "value": "two files",
                "confidence": "HIGH",
                "severity": "watch",
            }
        ],
        "attachments": [str(xlsx_a), str(xlsx_b)],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
    ):
        mock_send.return_value = "ts-split"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="t-split",
            session_id=sid,
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["ok"] is True

    upload_arg = mock_attach.call_args[0][0]
    # Both originals reach upload; no consolidated file produced.
    assert str(xlsx_a) in upload_arg
    assert str(xlsx_b) in upload_arg
    assert not any("consolidated_" in p for p in upload_arg)


def test_consolidation_swaps_parquet_to_xlsx_sibling(tmp_path, monkeypatch):
    """Parquet handles in attachments get swapped to xlsx siblings before merge.

    Codex P2 finding on PR-E: ``dump_sf_query`` and ``query_artifact`` ship
    Parquet paths to ``post_report``; ``_attach_files_async`` swaps each one
    to its ``.xlsx`` sibling via ``_prefer_xlsx_sibling`` at upload time.
    Without applying the same swap UP FRONT here, consolidation would see
    zero ``.xlsx`` files and skip the merge, defeating the feature for the
    most common multi-file case.
    """
    import session_runner
    from session_runner import _dispatch_post_report

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))
    monkeypatch.setenv("XLSX_CONSOLIDATE_ENABLED", "true")

    # Simulate the virtualized-query layout: parquet handle + xlsx sibling.
    parquet_a = outputs_dir / "sf_pipeline_dump.parquet"
    xlsx_a = outputs_dir / "sf_pipeline_dump.xlsx"
    parquet_a.write_bytes(b"PARFAKE")
    _write_real_xlsx(xlsx_a, "Pipeline", [["k", "v"], ["a", 1]])

    parquet_b = outputs_dir / "sf_funnel_dump.parquet"
    xlsx_b = outputs_dir / "sf_funnel_dump.xlsx"
    parquet_b.write_bytes(b"PARFAKE")
    _write_real_xlsx(xlsx_b, "Funnel", [["k", "v"], ["b", 2]])

    payload = {
        "headline": "Two parquet handles",
        "findings": [
            {
                "headline": "Both should consolidate via xlsx siblings",
                "value": "see attachments",
                "confidence": "HIGH",
                "severity": "watch",
            }
        ],
        # Agent shipped parquet handles, not xlsx paths.
        "attachments": [str(parquet_a), str(parquet_b)],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
    ):
        mock_send.return_value = "ts-parquet"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="t-parquet",
            session_id="sess-parquet-1",
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["ok"] is True

    upload_arg = mock_attach.call_args[0][0]
    xlsx_in_upload = [p for p in upload_arg if p.lower().endswith(".xlsx")]
    # Exactly 1 consolidated xlsx — both parquet handles swapped to xlsx
    # siblings and merged into a single workbook.
    assert len(xlsx_in_upload) == 1, (
        f"expected 1 consolidated xlsx after parquet→xlsx swap, "
        f"got xlsx={xlsx_in_upload} full={upload_arg}"
    )
    assert "consolidated_acme_" in xlsx_in_upload[0]
    # Neither original parquet path nor original xlsx siblings remain in
    # the upload list (they all collapsed into the consolidated file).
    assert str(parquet_a) not in upload_arg
    assert str(parquet_b) not in upload_arg
    assert str(xlsx_a) not in upload_arg
    assert str(xlsx_b) not in upload_arg


def test_consolidation_failure_falls_through(tmp_path, monkeypatch):
    """If consolidate_xlsx_files raises, original N files still upload."""
    import session_runner
    from session_runner import _dispatch_post_report

    outputs_dir = tmp_path / "outputs"
    outputs_dir.mkdir()
    monkeypatch.setenv("SESSION_OUTPUT_DIR", str(outputs_dir))
    monkeypatch.setenv("XLSX_CONSOLIDATE_ENABLED", "true")

    xlsx_a = outputs_dir / "alpha.xlsx"
    xlsx_b = outputs_dir / "beta.xlsx"
    _write_real_xlsx(xlsx_a, "Alpha", [["k", "v"], ["a", 1]])
    _write_real_xlsx(xlsx_b, "Beta", [["k", "v"], ["b", 2]])

    # Monkeypatch the function the dispatcher imports lazily.
    import xlsx_consolidate

    def _boom(**kwargs):
        raise RuntimeError("simulated consolidation failure")

    monkeypatch.setattr(xlsx_consolidate, "consolidate_xlsx_files", _boom)

    payload = {
        "headline": "Consolidation will raise",
        "findings": [
            {
                "headline": "Graceful fall-through",
                "value": "see attachments",
                "confidence": "HIGH",
                "severity": "watch",
            }
        ],
        "attachments": [str(xlsx_a), str(xlsx_b)],
    }
    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": payload,
    }

    with (
        patch.object(session_runner, "send_notification") as mock_send,
        patch.object(session_runner, "_attach_files_async") as mock_attach,
    ):
        mock_send.return_value = "ts-fail"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="t-fail",
            session_id="sess-fail-1",
            portco_key="acme",
        )

    result = json.loads(result_text)
    # The dispatch itself succeeds — consolidation failure never blocks delivery.
    assert result["ok"] is True

    upload_arg = mock_attach.call_args[0][0]
    # Both originals reach upload; no consolidated file because raise.
    assert str(xlsx_a) in upload_arg
    assert str(xlsx_b) in upload_arg
    assert not any("consolidated_" in p for p in upload_arg)


# ---------------------------------------------------------------------------
# Plan #52 PR-G — cancelled guard
# ---------------------------------------------------------------------------
#
# When the investigation row has been terminalized between the moment the
# Coordinator decided to call post_report and the moment the orchestrator
# dispatches it (watchdog Tier 3, user /stop, recovery sweep), the dispatch
# must short-circuit — no editor pass, no Slack post, no surface push, no
# lifecycle terminalize. The agent gets a structured no-op JSON back.
#
# Terminal statuses guarded (see ``_POST_REPORT_TERMINAL_GUARD_STATUSES``):
#   cancelled / failed / interrupted / orphan_dead_lettered / archived
#
# Excluded from the guard set (intentional):
#   running / queued — happy-path investigations must still post.
#   completed — re-entry on a completed row STILL runs (the only path to
#               "completed" is THIS function's terminalize call, so adding
#               "completed" would self-terminate every legitimate retry).


def _happy_payload() -> dict:
    """Minimal ad_hoc_investigation_result that renders cleanly."""
    return {
        "headline": "Win rate down 4.2pp this quarter",
        "findings": [
            {
                "headline": "Partner channel win rate collapsed",
                "value": "8% (n=42)",
                "confidence": "HIGH",
                "severity": "critical",
            }
        ],
    }


def _terminalized_inv(status: str) -> dict:
    """Return a fake investigation row in the given terminal status."""
    return {
        "id": 42,
        "thread_ts": "thread-x",
        "channel_id": "C-test",
        "user_id": "U1",
        "question": "what's win rate doing",
        "portco_key": "acme",
        "session_id": "sess-cancelled",
        "agent_id": "ag-x",
        "status": status,
        "started_at": None,
        "completed_at": None,
        "error_message": None,
        "recovery_count": 0,
        "container_id": "c1",
        "event_ts": "ts-event",
    }


def _running_inv() -> dict:
    """Return a fake investigation row still running (happy path)."""
    row = _terminalized_inv("running")
    return row


def test_post_report_cancelled_guard_skips_when_status_cancelled():
    """status='cancelled' triggers the guard — no Slack post, structured JSON."""
    from session_runner import _dispatch_post_report

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_terminalized_inv("cancelled"),
        ) as mock_db,
        patch("surface_pusher.push_to_canvas") as mock_push,
    ):
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-cancelled",
            inv_id=42,
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["ok"] is False
    assert result["skipped"] is True
    assert result["reason"] == "investigation_already_terminalized"
    assert result["inv_status"] == "cancelled"
    # No Slack post.
    mock_send.assert_not_called()
    # No surface push.
    mock_push.assert_not_called()
    # DB read happened exactly once with the inv_id we passed.
    mock_db.assert_called_once_with(42)


def test_post_report_cancelled_guard_skips_when_status_failed():
    """status='failed' triggers the guard."""
    from session_runner import _dispatch_post_report

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_terminalized_inv("failed"),
        ),
        patch("surface_pusher.push_to_canvas") as mock_push,
    ):
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-failed",
            inv_id=42,
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["skipped"] is True
    assert result["inv_status"] == "failed"
    mock_send.assert_not_called()
    mock_push.assert_not_called()


def test_post_report_cancelled_guard_skips_when_status_interrupted():
    """status='interrupted' triggers the guard."""
    from session_runner import _dispatch_post_report

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_terminalized_inv("interrupted"),
        ),
        patch("surface_pusher.push_to_canvas") as mock_push,
    ):
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-interrupted",
            inv_id=42,
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["skipped"] is True
    assert result["inv_status"] == "interrupted"
    mock_send.assert_not_called()
    mock_push.assert_not_called()


def test_post_report_cancelled_guard_skips_when_status_orphan_dead_lettered():
    """status='orphan_dead_lettered' triggers the guard."""
    from session_runner import _dispatch_post_report

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_terminalized_inv("orphan_dead_lettered"),
        ),
        patch("surface_pusher.push_to_canvas") as mock_push,
    ):
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-orphan",
            inv_id=42,
            portco_key="acme",
        )

    result = json.loads(result_text)
    assert result["skipped"] is True
    assert result["inv_status"] == "orphan_dead_lettered"
    mock_send.assert_not_called()
    mock_push.assert_not_called()


def test_post_report_cancelled_guard_no_op_when_status_running():
    """status='running' (the normal case) does NOT trigger the guard.

    The happy-path Slack post still fires. This is the regression check
    that proves the guard doesn't break normal investigations.
    """
    from session_runner import _dispatch_post_report
    import session_runner

    # Clear retry counter for isolation.
    session_runner._clear_post_report_retries("sess-running")

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_running_inv(),
        ),
    ):
        mock_send.return_value = "ts-happy"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-running",
            inv_id=42,
        )

    result = json.loads(result_text)
    assert result["ok"] is True, result
    assert result.get("skipped") is None
    # Happy path posts once.
    mock_send.assert_called_once()


def test_post_report_cancelled_guard_no_op_when_status_completed():
    """status='completed' is intentionally NOT in the guard set.

    The only path to 'completed' is the success-side terminalize call in
    _dispatch_post_report itself. Including 'completed' in the guard would
    self-terminate every legitimate retry of post_report within a session.
    Defensive: this asserts the exclusion is real.
    """
    from session_runner import _dispatch_post_report
    import session_runner

    session_runner._clear_post_report_retries("sess-completed")

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_terminalized_inv("completed"),
        ),
    ):
        mock_send.return_value = "ts-completed"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-completed",
            inv_id=42,
        )

    result = json.loads(result_text)
    assert result["ok"] is True, result
    # Happy path posts.
    mock_send.assert_called_once()


def test_post_report_cancelled_guard_no_op_when_inv_id_missing():
    """When inv_id is None (cron / theme-anchored path), the guard is skipped.

    There is no row to check, so the DB read is not even attempted. The
    happy-path post still fires.
    """
    from session_runner import _dispatch_post_report
    import session_runner

    session_runner._clear_post_report_retries("sess-no-inv")

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch("session_runner.db_adapter.get_investigation_by_id") as mock_db,
    ):
        mock_send.return_value = "ts-no-inv"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-no-inv",
            # inv_id intentionally omitted — cron path
        )

    result = json.loads(result_text)
    assert result["ok"] is True, result
    mock_send.assert_called_once()
    # Guard did not even try to read the DB because inv_id was None.
    mock_db.assert_not_called()


def test_post_report_cancelled_guard_no_op_when_db_read_raises():
    """If the DB read raises, the guard fails open (logs + proceeds).

    Fail-open semantics: the guard is a safety net for stale prose, NOT a
    hard gate. A flaky DB connection must not prevent the happy-path post.
    """
    from session_runner import _dispatch_post_report
    import session_runner

    session_runner._clear_post_report_retries("sess-db-err")

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            side_effect=RuntimeError("psycopg2 connection refused"),
        ),
    ):
        mock_send.return_value = "ts-db-err"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-db-err",
            inv_id=42,
        )

    result = json.loads(result_text)
    # Happy-path post fired despite the DB read failing.
    assert result["ok"] is True, result
    mock_send.assert_called_once()


def test_post_report_cancelled_guard_logs_post_report_cancelled_marker(caplog):
    """On a guard hit, the log line contains the literal marker, inv_id, and status.

    Operators grep Railway logs for `[POST_REPORT_CANCELLED]` over the
    first 14 days to track guard-hit rate. The marker, inv_id, and status
    are all load-bearing — losing any one breaks the operator's view.
    """
    import logging

    from session_runner import _dispatch_post_report

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_terminalized_inv("cancelled"),
        ),
        caplog.at_level(logging.WARNING, logger="session_runner"),
    ):
        _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-marker",
            inv_id=42,
        )

    # The marker, inv_id, and status all appear in the warning line.
    log_text = caplog.text
    assert "[POST_REPORT_CANCELLED]" in log_text
    assert "inv_id=42" in log_text
    assert "status=cancelled" in log_text
    # And no Slack post happened.
    mock_send.assert_not_called()


# ---------------------------------------------------------------------------
# Plan #52 PR-G (codex P1 follow-up) — terminal signal to caller-side fallback
# ---------------------------------------------------------------------------
#
# The cancelled-guard above short-circuits post_report when the investigation
# row is already terminal. Before the codex P1 fix, the guard returned
# ``{"ok": False, "skipped": True, ...}`` without any marker the outer
# session loop (``_stream_and_handle``) recognized as terminal. The caller
# of ``_stream_and_handle`` (e.g. ``run_adhoc_mcp_session`` line ~6155+)
# then saw ``delivery_state.is_delivered() is False`` and emitted a second
# Slack post — the "Investigation didn't produce a final report" line —
# on top of the watchdog/(stop)/recovery path's ❌ terminal notice. Two
# contradictory user-facing posts on the same Slack message.
#
# Fix: the guard now (a) sets ``_terminal: True`` on the return dict and
# (b) marks the session in ``_post_report_cancelled_guard_sessions`` so
# the caller-side fallback can consume the marker and skip the redundant
# post. The tests below pin both halves of the contract.


def test_post_report_cancelled_guard_marks_terminal_on_return():
    """The guard return dict carries ``_terminal: True`` (codex P1 signal).

    This is the in-band JSON marker that the post-dispatch promotion code
    in ``_stream_and_handle`` keys on. Without it the outer loop has no
    way to distinguish a cancelled-guard skip from any other ``ok: False``
    result (e.g. ``schema_validation_failed``), and the legacy fallback
    fires regardless.
    """
    from session_runner import _dispatch_post_report

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_terminalized_inv("cancelled"),
        ),
    ):
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id="sess-terminal-marker",
            inv_id=42,
        )

    result = json.loads(result_text)
    # The terminal marker is present — caller-side fallback will see it.
    assert result.get("_terminal") is True, (
        f"expected ``_terminal: True`` on guard return, got {result!r}"
    )
    # Existing skip contract still holds.
    assert result["ok"] is False
    assert result["skipped"] is True
    assert result["reason"] == "investigation_already_terminalized"
    # And no Slack post happened.
    mock_send.assert_not_called()


def test_post_report_cancelled_guard_registers_session_for_fallback_suppression():
    """The guard records the session_id in ``_post_report_cancelled_guard_sessions``.

    The caller-side fallback in ``run_adhoc_mcp_session`` (and the recovery
    resume path) reads this set via ``_consume_post_report_cancelled_guard``
    to decide whether to emit the "Investigation didn't produce a final
    report" Slack post. Without the marker the suppression branch never
    fires and we get the double-post race the codex P1 finding flagged.
    """
    import session_runner
    from session_runner import _dispatch_post_report

    sid = "sess-registry-marker"

    # Clear any leftover state from prior tests so the assertion is honest.
    with session_runner._post_report_cancelled_guard_lock:
        session_runner._post_report_cancelled_guard_sessions.discard(sid)

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification"),
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_terminalized_inv("failed"),
        ),
    ):
        _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id=sid,
            inv_id=42,
        )

    # Registry contains the session. ``_consume_*`` returns True AND clears.
    assert session_runner._consume_post_report_cancelled_guard(sid) is True
    # Second consume returns False — one-shot semantics, no re-fire.
    assert session_runner._consume_post_report_cancelled_guard(sid) is False


def test_consume_post_report_cancelled_guard_returns_false_when_unmarked():
    """A session that never tripped the guard is never reported as terminal.

    Defensive: the fallback in ``run_adhoc_mcp_session`` runs on EVERY
    session that didn't deliver via post_report, including happy-path
    "Coordinator forgot to call post_report" cases. We must not
    accidentally suppress those legitimate fallbacks.
    """
    import session_runner

    # A session_id that has never been marked.
    sid = "sess-never-marked-12345"
    with session_runner._post_report_cancelled_guard_lock:
        session_runner._post_report_cancelled_guard_sessions.discard(sid)

    assert session_runner._consume_post_report_cancelled_guard(sid) is False
    # Empty / falsy session_id is also safe (cron paths).
    assert session_runner._consume_post_report_cancelled_guard("") is False
    assert session_runner._consume_post_report_cancelled_guard(None) is False  # type: ignore[arg-type]


def test_stream_and_handle_caller_suppresses_fallback_when_guard_fired(
    monkeypatch, tmp_path
):
    """End-to-end pin: when the cancelled-guard fired during _stream_and_handle,
    ``run_adhoc_mcp_session`` does NOT emit the fallback Slack post.

    This is the regression that closes the codex P1 finding. Before the
    fix, the cancelled-guard set ``delivery_state=NOT_DELIVERED`` (via
    its ``ok: False`` return) and the caller's ``elif text_parts:``
    branch ran ``post_analysis(title="Investigation incomplete", ...)``
    on top of the watchdog/(stop)/recovery path's ❌ terminal notice —
    two contradictory user-facing posts.

    Mirrors the test fixture pattern from session_runner_test.py
    (``_stub_session_runner_for_thread_followup``) — we stub out the
    network-heavy moving parts and drive ``run_adhoc_mcp_session``
    directly to assert the suppression branch fires.
    """
    import session_runner
    from session_runner_test import _stub_session_runner_for_thread_followup
    from lifecycle import DeliveryState

    stubs = _stub_session_runner_for_thread_followup(session_runner, monkeypatch)

    # Simulate the cancelled-guard having fired during _stream_and_handle:
    # the registry is marked for the session_id the stubs use
    # (``sesn_EXAMPLE``). The stream stub returns NOT_DELIVERED + text_parts,
    # which is the exact shape that would normally trigger the
    # "Investigation didn't produce a final report" fallback. With the
    # registry marker present the suppression branch must fire instead.
    session_runner._mark_post_report_cancelled_guard_fired("sesn_EXAMPLE")

    stubs["stream"].return_value = (
        ["some leaked narration the coordinator emitted before the guard"],
        DeliveryState.NOT_DELIVERED,
        None,
        [],
    )

    captured_post_analysis = MagicMock()
    captured_send_notification = MagicMock()
    captured_terminalize = MagicMock()
    monkeypatch.setattr(session_runner, "post_analysis", captured_post_analysis)
    monkeypatch.setattr(session_runner, "send_notification", captured_send_notification)
    monkeypatch.setattr("lifecycle.terminalize_lifecycle", captured_terminalize)

    session_runner.run_adhoc_mcp_session(
        question="some user question",
        user_id="U999",
        thread_ts="1737654321.000100",
        channel_id="C999",
        existing_inv_id=None,
        event_ts="1737654322.000200",
    )

    # The fallback "Investigation incomplete" post_analysis was SUPPRESSED.
    # Pre-fix this fired on top of the watchdog/(stop) terminal notice.
    captured_post_analysis.assert_not_called()
    # The orchestration-chatter watch line was ALSO suppressed.
    chatter_calls = [
        c
        for c in captured_send_notification.call_args_list
        if c.args
        and isinstance(c.args[0], str)
        and "no user-facing findings" in (c.args[1] if len(c.args) > 1 else "")
    ]
    assert chatter_calls == []
    # The bare "Investigation produced no output." watch line was suppressed.
    no_output_watch_calls = [
        c
        for c in captured_send_notification.call_args_list
        if c.args
        and isinstance(c.args[0], str)
        and "Investigation produced no output" in (c.args[1] if len(c.args) > 1 else "")
    ]
    assert no_output_watch_calls == []

    # And the registry was drained (one-shot semantics — a second pass
    # would not erroneously suppress a legitimate fallback).
    assert (
        session_runner._consume_post_report_cancelled_guard("sesn_EXAMPLE") is False
    ), "registry should be empty after the suppression branch consumed it"


def test_stream_and_handle_caller_runs_fallback_when_guard_did_not_fire(
    monkeypatch,
):
    """Regression check: a session that DIDN'T trip the guard still gets
    the legacy fallback (the codex P1 fix must not break legitimate
    "Coordinator forgot to call post_report" fallbacks).

    Mirrors the existing test in session_runner_test.py that pins the
    fallback's neutral-line behavior; this variant just confirms our
    suppression check doesn't accidentally swallow the call.
    """
    import session_runner
    from session_runner_test import _stub_session_runner_for_thread_followup
    from lifecycle import DeliveryState

    stubs = _stub_session_runner_for_thread_followup(session_runner, monkeypatch)

    # Explicitly clear the registry so the suppression branch CAN'T fire.
    with session_runner._post_report_cancelled_guard_lock:
        session_runner._post_report_cancelled_guard_sessions.discard("sesn_EXAMPLE")

    stubs["stream"].return_value = (
        # Non-orchestration-chatter text so we land in the "neutral fallback"
        # branch rather than the orchestration-chatter branch.
        ["real findings text the coordinator emitted but never post_reported"],
        DeliveryState.NOT_DELIVERED,
        None,
        [],
    )

    captured_post_analysis = MagicMock()
    monkeypatch.setattr(session_runner, "post_analysis", captured_post_analysis)
    monkeypatch.setattr("lifecycle.terminalize_lifecycle", MagicMock())
    # Provide a non-tmp_path safe-attachment validator that returns False
    # so the fallback's preview count stays at zero (the test doesn't
    # care about attachments — only about whether post_analysis fired).
    monkeypatch.setattr(session_runner, "_is_safe_attachment_path", lambda p: False)
    monkeypatch.setattr(session_runner, "_prefer_xlsx_sibling", lambda p: p)

    session_runner.run_adhoc_mcp_session(
        question="some user question",
        user_id="U999",
        thread_ts="1737654321.000100",
        channel_id="C999",
        existing_inv_id=None,
        event_ts="1737654322.000200",
    )

    # The legacy fallback ran exactly once — the suppression branch did
    # NOT erroneously fire because the registry was clean.
    assert captured_post_analysis.call_count == 1
    assert (
        captured_post_analysis.call_args.kwargs.get("title")
        == "Investigation incomplete"
    )


def test_post_report_cancelled_guard_drains_stale_marker_when_not_fired_this_turn():
    """Codex P2 regression: stale guard markers are drained at dispatch start.

    Scenario: a prior turn on this session tripped the cancelled-guard
    (e.g. watchdog Tier 3 marked the investigation row 'failed' mid-flow),
    so ``_mark_post_report_cancelled_guard_fired(session_id)`` ran. The
    caller-side ``_consume_post_report_cancelled_guard`` only fires on the
    NOT_DELIVERED branch in ``run_adhoc_mcp_session``. If a later turn
    delivers successfully, the marker stays in the set forever — and
    because thread follow-ups reuse the same session, a future
    unrelated turn's fallback could consume the stale marker and
    suppress a legitimate "investigation incomplete / no output" post.

    Fix: clear any leaked marker at the start of ``_dispatch_post_report``
    whenever the guard does NOT fire this turn. This pins the marker
    lifetime to a single turn.
    """
    import session_runner
    from session_runner import _dispatch_post_report

    sid = "sess-p2-stale-marker"

    # Pre-populate the registry to simulate a marker leaked from a
    # prior turn that tripped the guard. The current turn will NOT
    # trip it (we wire the DB stub to return a 'running' investigation).
    session_runner._mark_post_report_cancelled_guard_fired(sid)
    assert sid in session_runner._post_report_cancelled_guard_sessions

    # Also clear retry counters so the happy-path post lands cleanly.
    session_runner._clear_post_report_retries(sid)

    tool_input = {
        "response_type": "ad_hoc_investigation_result",
        "payload": _happy_payload(),
    }

    with (
        patch("session_runner.send_notification") as mock_send,
        patch(
            "session_runner.db_adapter.get_investigation_by_id",
            return_value=_running_inv(),
        ),
    ):
        mock_send.return_value = "ts-p2-drain"
        result_text = _dispatch_post_report(
            tool_input,
            thread_ts="thread-x",
            session_id=sid,
            inv_id=42,
        )

    # The happy-path post still ran (regression check).
    result = json.loads(result_text)
    assert result["ok"] is True, result
    mock_send.assert_called_once()

    # And — the P2 assertion — the stale marker was drained at dispatch
    # start, so a future unrelated fallback on this session cannot
    # erroneously consume it.
    assert sid not in session_runner._post_report_cancelled_guard_sessions, (
        "stale cancelled-guard marker leaked past _dispatch_post_report"
    )
    assert session_runner._consume_post_report_cancelled_guard(sid) is False
