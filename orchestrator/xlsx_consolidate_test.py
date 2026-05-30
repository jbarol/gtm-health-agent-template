"""Tests for orchestrator/xlsx_consolidate.consolidate_xlsx_files.

Covers Plan #48 §8 (a, c, d, e, f) plus the banner-row acceptance
criterion D5 and the empty-list edge case. The split-files registry
test (b) is deferred to Phase 2 along with the session_runner wiring.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

ORCH = Path(__file__).resolve().parent
if str(ORCH) not in sys.path:
    sys.path.insert(0, str(ORCH))


def _write_simple_xlsx(path: Path, rows: int = 5) -> str:
    """Write an xlsx with a header + ``rows`` data rows. Returns str(path)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["Id", "Name", "Value"])
    for i in range(rows):
        ws.append([f"id-{i}", f"name-{i}", i * 10])
    wb.save(str(path))
    return str(path)


def _write_wide_xlsx(path: Path, rows: int, cols: int = 20) -> str:
    """Write a synthetic xlsx large enough to test the size cap."""
    wb = Workbook()
    ws = wb.active
    ws.append([f"col_{c}" for c in range(cols)])
    payload = "x" * 80  # ~80 bytes per cell to inflate file size cheaply
    for i in range(rows):
        ws.append([f"{payload}_{i}_{c}" for c in range(cols)])
    wb.save(str(path))
    return str(path)


def test_consolidation_merges_multiple_xlsx_into_one(tmp_path):
    """Three xlsx files → one consolidated workbook with three sheets."""
    from xlsx_consolidate import consolidate_xlsx_files

    paths = [
        _write_simple_xlsx(
            tmp_path / f"sf_data_{i}_20260518T065023000000_abc{i}.xlsx", rows=5
        )
        for i in range(3)
    ]
    merged, did = consolidate_xlsx_files(paths, output_dir=str(tmp_path))
    assert did is True
    assert len(merged) == 1
    assert merged[0].endswith(".xlsx")
    assert os.path.exists(merged[0])

    wb = load_workbook(merged[0], read_only=True)
    try:
        assert len(wb.sheetnames) == 3
    finally:
        wb.close()


def test_consolidation_includes_banner_row(tmp_path):
    """Every consolidated sheet has 'Source: <basename>' at row 1 (D5)."""
    from xlsx_consolidate import consolidate_xlsx_files

    p1 = _write_simple_xlsx(
        tmp_path / "sf_alpha_20260518T065023000000_a3f1.xlsx", rows=2
    )
    p2 = _write_simple_xlsx(
        tmp_path / "sf_beta_20260518T070001000000_b4c2.xlsx", rows=2
    )
    merged, did = consolidate_xlsx_files([p1, p2], output_dir=str(tmp_path))
    assert did is True

    wb = load_workbook(merged[0], read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # Row 1: banner. Row 2: headers. Row 3+: data.
            banner = rows[0]
            assert banner[0] is not None
            assert banner[0].startswith("Source: ")
            assert banner[0].endswith(".xlsx")
            # Row 2 should be the original header row.
            assert rows[1] == ("Id", "Name", "Value")
            # Row 3 onward is data.
            assert rows[2][0] == "id-0"
    finally:
        wb.close()


def test_size_cap_fallback_returns_original_paths(tmp_path):
    """Files totaling > cap → original paths returned, no consolidated file."""
    from xlsx_consolidate import consolidate_xlsx_files

    p1 = _write_wide_xlsx(
        tmp_path / "sf_big1_20260518T065023000000_aaa1.xlsx", rows=200
    )
    p2 = _write_wide_xlsx(
        tmp_path / "sf_big2_20260518T070001000000_bbb2.xlsx", rows=200
    )
    # Force the cap to fire: both files together exceed 0.001 MB easily.
    merged, did = consolidate_xlsx_files([p1, p2], str(tmp_path), size_cap_mb=0.001)
    assert did is False
    assert set(merged) == {p1, p2}
    # No consolidated file should have been written.
    assert not any(f.startswith("consolidated_") for f in os.listdir(tmp_path))


def test_sheet_name_collision_deduped(tmp_path):
    """Two xlsx files with the same derived sheet name → dedup suffix appended."""
    from xlsx_consolidate import consolidate_xlsx_files

    p1 = _write_simple_xlsx(
        tmp_path / "sf_leads_20260518T065023000000_a3f1.xlsx", rows=3
    )
    p2 = _write_simple_xlsx(
        tmp_path / "sf_leads_20260518T070001000000_b4c2.xlsx", rows=3
    )
    merged, did = consolidate_xlsx_files([p1, p2], str(tmp_path))
    assert did is True

    wb = load_workbook(merged[0], read_only=True)
    try:
        names = wb.sheetnames
        assert len(set(names)) == 2  # no duplicates
        # The second sheet should carry a " 2" dedup suffix.
        assert any(n.endswith(" 2") for n in names)
    finally:
        wb.close()


def test_sheet_name_invalid_chars_sanitized():
    """Chars in :\\/?*[] are replaced with underscores; result ≤31 chars."""
    from xlsx_consolidate import _sanitize_sheet_name

    raw = "Report: Q1/Q2 [vs prior]"
    safe = _sanitize_sheet_name(raw)
    assert not any(c in safe for c in r":\/?*[]")
    assert len(safe) <= 31


def test_sheet_name_truncated_long_input():
    """Stem longer than 28 chars truncates with room for the dedup suffix."""
    from xlsx_consolidate import _sanitize_sheet_name

    raw = "a_really_long_sheet_name_that_exceeds_excel_31_char_limit"
    safe = _sanitize_sheet_name(raw)
    assert len(safe) <= 28  # leaves room for " 2"/" 3" suffix


def test_single_xlsx_skips_consolidation(tmp_path):
    """One xlsx → return as-is, did_consolidate=False."""
    from xlsx_consolidate import consolidate_xlsx_files

    p = _write_simple_xlsx(
        tmp_path / "sf_opps_20260518T065023000000_c5d3.xlsx", rows=10
    )
    merged, did = consolidate_xlsx_files([p], str(tmp_path))
    assert did is False
    assert merged == [p]


def test_empty_input_returns_empty_no_op(tmp_path):
    """Zero-path input → ([], False); never raises, never writes a file."""
    from xlsx_consolidate import consolidate_xlsx_files

    merged, did = consolidate_xlsx_files([], str(tmp_path))
    assert merged == []
    assert did is False
    assert os.listdir(tmp_path) == []


def test_coordinator_produced_file_sorts_first(tmp_path):
    """Coordinator-named file appears as the first sheet."""
    from xlsx_consolidate import consolidate_xlsx_files

    # Coordinator-named (no sf_ prefix + ts/hex pattern).
    coord = _write_simple_xlsx(tmp_path / "call_prep_brief.xlsx", rows=2)
    # SF dump.
    dump = _write_simple_xlsx(
        tmp_path / "sf_pipeline_opps_20260518T065023000000_a3f1.xlsx", rows=2
    )

    # Pass in reverse order; sort should put coordinator first.
    merged, did = consolidate_xlsx_files([dump, coord], str(tmp_path))
    assert did is True

    wb = load_workbook(merged[0], read_only=True)
    try:
        # The first sheet should derive from the coordinator-produced stem.
        assert wb.sheetnames[0].lower().startswith("call prep")
    finally:
        wb.close()


def test_failure_returns_original_paths(tmp_path, monkeypatch):
    """Any exception in the merge loop → (original_paths, False), no raise."""
    from xlsx_consolidate import consolidate_xlsx_files

    p1 = _write_simple_xlsx(tmp_path / "sf_a_20260518T065023000000_a3f1.xlsx", rows=2)
    p2 = _write_simple_xlsx(tmp_path / "sf_b_20260518T070001000000_b4c2.xlsx", rows=2)

    import openpyxl

    def _boom(*args, **kwargs):
        raise IOError("simulated openpyxl failure")

    monkeypatch.setattr(openpyxl, "load_workbook", _boom)

    merged, did = consolidate_xlsx_files([p1, p2], str(tmp_path))
    assert did is False
    assert set(merged) == {p1, p2}


def _write_multi_sheet_xlsx(path: Path, sheet_names: list[str]) -> str:
    """Write an xlsx with N named sheets, each with a header + 2 rows."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    for name in sheet_names:
        ws = wb.create_sheet(title=name)
        ws.append(["Id", "Name"])
        ws.append([1, f"{name}-row1"])
        ws.append([2, f"{name}-row2"])
    wb.save(str(path))
    return str(path)


def test_multi_sheet_source_preserved(tmp_path):
    """Codex P2 fix: every sheet from a multi-sheet source survives consolidation.

    Before the fix, only ``worksheets[0]`` was copied — sheets 2..N were
    silently dropped from the consolidated output, losing datasets the
    Coordinator had explicitly grouped via ``materialize_xlsx(sheets=[...])``.
    """
    from xlsx_consolidate import consolidate_xlsx_files

    # Multi-sheet coordinator output + a single-sheet sf dump
    multi = _write_multi_sheet_xlsx(
        tmp_path / "pipeline_briefing.xlsx",
        ["summary", "by_sector", "renewal_risk"],
    )
    single = _write_simple_xlsx(
        tmp_path / "sf_opps_20260518T065023000000_a3f1.xlsx", rows=2
    )
    merged, did = consolidate_xlsx_files([multi, single], str(tmp_path))
    assert did is True

    wb = load_workbook(merged[0], read_only=True)
    try:
        names = wb.sheetnames
        # Multi-sheet source contributes 3 sheets; single source contributes 1
        assert len(names) == 4
        # Verify each source sheet shows up (sanitized title, possibly deduped)
        normalized = [n.lower() for n in names]
        assert any("summary" in n for n in normalized)
        assert any("by_sector" in n or "by sector" in n for n in normalized)
        assert any("renewal" in n for n in normalized)
    finally:
        wb.close()


def test_multi_sheet_name_collision_deduped(tmp_path):
    """Two multi-sheet sources with overlapping internal titles still produce
    unique sheet names in the consolidated workbook."""
    from xlsx_consolidate import consolidate_xlsx_files

    a = _write_multi_sheet_xlsx(tmp_path / "report_a.xlsx", ["summary", "detail"])
    b = _write_multi_sheet_xlsx(tmp_path / "report_b.xlsx", ["summary", "detail"])
    merged, did = consolidate_xlsx_files([a, b], str(tmp_path))
    assert did is True

    wb = load_workbook(merged[0], read_only=True)
    try:
        names = wb.sheetnames
        assert len(names) == 4
        assert len(set(names)) == 4  # all unique
        # Dedup suffix should appear on the collisions
        assert any(n.endswith(" 2") for n in names)
    finally:
        wb.close()


def test_output_path_unique_per_call(tmp_path):
    """Codex P2 fix: two consecutive calls produce different output filenames.

    Before the fix, the deterministic ``consolidated_<portco>_<date>.xlsx``
    name caused a second concurrent investigation in the same portco/date
    to overwrite the first call's workbook before its async Slack upload
    could read it. The microsecond timestamp + random suffix prevents this.
    """
    from xlsx_consolidate import consolidate_xlsx_files

    paths = [
        _write_simple_xlsx(
            tmp_path / f"sf_data_{i}_20260518T065023000000_abc{i}.xlsx", rows=3
        )
        for i in range(2)
    ]

    merged_a, did_a = consolidate_xlsx_files(
        paths, str(tmp_path), portco_key="acme"
    )
    merged_b, did_b = consolidate_xlsx_files(
        paths, str(tmp_path), portco_key="acme"
    )
    assert did_a is True
    assert did_b is True

    # Distinct output paths — second call did not overwrite the first.
    assert merged_a[0] != merged_b[0]
    assert os.path.exists(merged_a[0])
    assert os.path.exists(merged_b[0])
    # Both follow the new naming convention.
    for path in (merged_a[0], merged_b[0]):
        assert os.path.basename(path).startswith("consolidated_acme_")
        assert path.endswith(".xlsx")


# ---------------------------------------------------------------------------
# Plan #52 PR-D: split-files registry tests
# Registry lives in session_runner; tested here to centralize Plan #48 coverage.
# ---------------------------------------------------------------------------


def test_split_files_keyword_registers_preference():
    """'separate files' in question → split preference is True."""
    import sys as _sys
    from pathlib import Path as _Path

    _sys.path.insert(0, str(_Path(__file__).parent))
    from session_runner import (
        _detect_split_files,
        _register_split_files_pref,
        _consume_split_files_pref,
    )

    _register_split_files_pref(
        "sess-kw-1", _detect_split_files("give me separate files")
    )
    assert _consume_split_files_pref("sess-kw-1") is True


def test_no_keyword_gives_false():
    import sys as _sys
    from pathlib import Path as _Path

    _sys.path.insert(0, str(_Path(__file__).parent))
    from session_runner import _detect_split_files

    assert _detect_split_files("show me Q1 pipeline by stage") is False


def test_consume_returns_false_when_unregistered():
    import sys as _sys
    from pathlib import Path as _Path

    _sys.path.insert(0, str(_Path(__file__).parent))
    from session_runner import _consume_split_files_pref

    # No registration → default False (consolidate)
    assert _consume_split_files_pref("sess-never-registered") is False


def test_split_pref_cleared_on_consume():
    import sys as _sys
    from pathlib import Path as _Path

    _sys.path.insert(0, str(_Path(__file__).parent))
    from session_runner import _register_split_files_pref, _consume_split_files_pref

    _register_split_files_pref("sess-clear-1", True)
    assert _consume_split_files_pref("sess-clear-1") is True
    # Second consume: cleared, returns False
    assert _consume_split_files_pref("sess-clear-1") is False
